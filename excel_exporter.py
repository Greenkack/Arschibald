"""excel_exporter.py - Excel Export System"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, Any, List, Optional
from datetime import datetime

class ExcelExporter:
    """Excel-Export-Manager mit Styling"""
    
    def __init__(self):
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def export_dataframe(self, df: pd.DataFrame, file_path: str, sheet_name: str = "Data", 
                        apply_styling: bool = True):
        """Exportiere DataFrame nach Excel mit Styling"""
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Füge Daten hinzu
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Styling
                if apply_styling:
                    if r_idx == 1:  # Header
                        cell.font = self.header_font
                        cell.fill = self.header_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    cell.border = self.border
        
        # Auto-width
        if apply_styling:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(file_path)
    
    def export_calculation_results(self, results: Dict[str, Any], file_path: str):
        """Exportiere Berechnungsergebnisse"""
        # Erstelle DataFrames für verschiedene Bereiche
        summary_data = []
        for key, value in results.items():
            if isinstance(value, (int, float, str)):
                summary_data.append({'Parameter': key, 'Wert': value})
        
        df_summary = pd.DataFrame(summary_data)
        
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df_summary.to_excel(writer, sheet_name='Zusammenfassung', index=False)
            
            # Falls es Listen/Arrays gibt
            if 'monthly_values' in results and isinstance(results['monthly_values'], list):
                df_monthly = pd.DataFrame({
                    'Monat': range(1, len(results['monthly_values']) + 1),
                    'Wert': results['monthly_values']
                })
                df_monthly.to_excel(writer, sheet_name='Monatswerte', index=False)
    
    def export_multi_sheet(self, data_dict: Dict[str, pd.DataFrame], file_path: str):
        """Exportiere mehrere Sheets"""
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            for sheet_name, df in data_dict.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    def create_report(self, title: str, data: pd.DataFrame, file_path: str, 
                     metadata: Optional[Dict[str, str]] = None):
        """Erstelle formatierten Report"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"
        
        # Titel
        ws['A1'] = title
        ws['A1'].font = Font(size=16, bold=True)
        
        # Metadata
        if metadata:
            row = 3
            for key, value in metadata.items():
                ws.cell(row=row, column=1, value=key)
                ws.cell(row=row, column=2, value=value)
                row += 1
        
        # Daten ab Zeile 7
        start_row = 7
        for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), start_row):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        wb.save(file_path)
