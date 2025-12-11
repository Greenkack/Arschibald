"""
Controlling System Report Generator

Provides report generation, saving, loading, listing, and export functionality
for the Employee Controlling System.

Requirements: 9.1, 9.2, 13.2, 13.3, 13.5, 14.1, 14.2, 14.3, 14.4, 15.1, 15.2,
              18.1, 20.1, 20.2
"""

import sys
import logging
import json
import io
from pathlib import Path
from datetime import date, datetime, timedelta
from typing import List, Dict, Any, Optional
from sqlalchemy.orm import Session

# PDF export
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image,
        KeepTogether, PageBreak
    )
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# Chart to image conversion
try:
    import plotly.io as pio
    PLOTLY_AVAILABLE = True
except ImportError:
    PLOTLY_AVAILABLE = False

# Excel export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Add project root to path
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from controlling.models import (  # noqa: E402
    Employee,
    Report,
    ReportType
)
from controlling.analytics import AnalyticsEngine  # noqa: E402
from controlling.team_analytics import TeamAnalytics  # noqa: E402
from controlling.pdf_config import get_color_scheme  # noqa: E402

logger = logging.getLogger(__name__)


# Deutsche Monatsnamen
GERMAN_MONTHS = {
    1: "Januar", 2: "Februar", 3: "März", 4: "April",
    5: "Mai", 6: "Juni", 7: "Juli", 8: "August",
    9: "September", 10: "Oktober", 11: "November", 12: "Dezember"
}

# Deutsche Berichtstyp-Namen
REPORT_TYPE_NAMES = {
    "DAILY": "Täglich",
    "WEEKLY": "Wöchentlich",
    "MONTHLY": "Monatlich",
    "QUARTERLY": "Vierteljährlich",
    "YEARLY": "Jährlich",
    "SINCE_START": "Seit Arbeitsbeginn"
}


def format_date_german(date_obj: date) -> str:
    """
    Formatiere Datum im deutschen Format: '01. Januar 2025'
    
    Args:
        date_obj: Date-Objekt
        
    Returns:
        Formatierter String
    """
    return f"{date_obj.day:02d}. {GERMAN_MONTHS[date_obj.month]} {date_obj.year}"


def format_datetime_german(datetime_str: str) -> str:
    """
    Formatiere ISO-DateTime-String zu deutschem Datum: '12. Dezember 2025'
    
    Args:
        datetime_str: ISO DateTime String (z.B. '2025-12-07T02:13:19.618798')
        
    Returns:
        Formatierter String
    """
    try:
        dt = datetime.fromisoformat(datetime_str)
        return format_date_german(dt.date())
    except:
        return datetime_str


def get_report_type_name_german(report_type: str) -> str:
    """
    Hole deutschen Namen für Berichtstyp.
    
    Args:
        report_type: Berichtstyp-Key (z.B. 'SINCE_START')
        
    Returns:
        Deutscher Name
    """
    return REPORT_TYPE_NAMES.get(report_type, report_type)


class ReportGenerator:
    """
    Report generator for creating, saving, and loading performance reports.

    Requirements: 9.1, 9.2, 13.2, 13.3, 13.5, 15.1, 15.2, 18.1, 20.1, 20.2
    """

    def __init__(self, db: Session):
        self.db = db
        self.analytics = AnalyticsEngine(db)
        self.team_analytics = TeamAnalytics(db)

    def _calculate_date_range(
        self,
        report_type: ReportType,
        reference_date: Optional[date] = None,
        employee_start_date: Optional[date] = None
    ) -> tuple[date, date]:
        """
        Calculate start and end dates for a report based on report type.

        Args:
            report_type: Type of report (DAILY, WEEKLY, MONTHLY, etc.)
            reference_date: Reference date for the report (defaults to today)
            employee_start_date: Employee's start date
                (required for SINCE_START)

        Returns:
            Tuple of (start_date, end_date)
        """
        if reference_date is None:
            reference_date = date.today()

        if report_type == ReportType.DAILY:
            return reference_date, reference_date

        elif report_type == ReportType.WEEKLY:
            # Monday of the week
            start = reference_date - timedelta(days=reference_date.weekday())
            # Sunday of the week
            end = start + timedelta(days=6)
            return start, end

        elif report_type == ReportType.MONTHLY:
            # First day of month
            start = reference_date.replace(day=1)
            # Last day of month
            if reference_date.month == 12:
                end = reference_date.replace(day=31)
            else:
                next_month = reference_date.replace(
                    month=reference_date.month + 1, day=1
                )
                end = next_month - timedelta(days=1)
            return start, end

        elif report_type == ReportType.QUARTERLY:
            # Determine quarter
            quarter = (reference_date.month - 1) // 3 + 1
            start_month = (quarter - 1) * 3 + 1
            start = reference_date.replace(month=start_month, day=1)

            # Last day of quarter
            end_month = quarter * 3
            if end_month == 12:
                end = reference_date.replace(month=12, day=31)
            else:
                next_quarter = reference_date.replace(
                    month=end_month + 1, day=1
                )
                end = next_quarter - timedelta(days=1)
            return start, end

        elif report_type == ReportType.YEARLY:
            start = reference_date.replace(month=1, day=1)
            end = reference_date.replace(month=12, day=31)
            return start, end

        elif report_type == ReportType.SINCE_START:
            if employee_start_date is None:
                raise ValueError(
                    "employee_start_date required for "
                    "SINCE_START reports"
                )
            return employee_start_date, date.today()

        else:
            raise ValueError(f"Unknown report type: {report_type}")

    def generate_report(
        self,
        employee_id: int,
        report_type: ReportType,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None
    ) -> Dict[str, Any]:
        """
        Generate a performance report for a single employee.

        Args:
            employee_id: ID of the employee
            report_type: Type of report to generate
            start_date: Optional custom start date
            end_date: Optional custom end date

        Returns:
            Dictionary containing report data

        Requirements: 9.1, 9.2
        """
        # Get employee
        employee = self.db.query(Employee).filter(
            Employee.id == employee_id
        ).first()

        if not employee:
            raise ValueError(f"Employee with ID {employee_id} not found")

        # Calculate date range if not provided
        if start_date is None or end_date is None:
            start_date, end_date = self._calculate_date_range(
                report_type,
                reference_date=end_date or date.today(),
                employee_start_date=employee.start_date
            )

        # Get aggregated data from analytics engine
        aggregated_data = self.analytics.aggregate_data(
            employee_id,
            report_type,
            start_date,
            end_date
        )

        # Extract quotas and ratios from aggregated data
        quotas = aggregated_data.get("quotas", {})
        ratio_descriptions = aggregated_data.get("ratios", {})

        # Get team information
        team_name = None
        team_leader = None
        if hasattr(employee, 'team') and employee.team:
            team_name = employee.team.name
            if employee.team.team_leader_id:
                try:
                    from controlling.employee_manager import EmployeeManager
                    emp_manager = EmployeeManager(self.db)
                    leader = emp_manager.get_employee(employee.team.team_leader_id)
                    if leader:
                        team_leader = leader.display_name
                except Exception:
                    pass

        # Build report data
        report_data = {
            "employee_id": employee_id,
            "employee_name": employee.full_name,
            "agent_name": employee.agent_name,
            "position": employee.position.name if employee.position else None,
            "team_name": team_name,
            "team_leader": team_leader,
            "report_type": report_type.value,
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "aggregated_data": aggregated_data,
            "quotas": quotas,
            "ratio_descriptions": ratio_descriptions,
            "generated_at": datetime.now().isoformat()
        }

        return report_data

    def generate_comparison_report(
        self,
        employee_ids: List[int],
        report_type: ReportType,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None
    ) -> Dict[str, Any]:
        """
        Generate a comparison report for multiple employees.

        Args:
            employee_ids: List of employee IDs (max 10)
            report_type: Type of report to generate
            start_date: Optional custom start date
            end_date: Optional custom end date

        Returns:
            Dictionary containing comparison report data

        Requirements: 20.1, 20.2
        """
        if len(employee_ids) > 10:
            raise ValueError(
                "Comparison reports support maximum 10 employees"
            )

        if len(employee_ids) < 2:
            raise ValueError(
                "Comparison reports require at least 2 employees"
            )

        # Generate individual reports for each employee
        employee_reports = []
        for emp_id in employee_ids:
            try:
                report = self.generate_report(
                    emp_id,
                    report_type,
                    start_date,
                    end_date
                )
                employee_reports.append(report)
            except ValueError as e:
                logger.warning(f"Skipping employee {emp_id}: {e}")
                continue

        if not employee_reports:
            raise ValueError("No valid employees found for comparison")

        # Use date range from first report
        first_report = employee_reports[0]

        # Build comparison report
        comparison_data = {
            "report_type": report_type.value,
            "start_date": first_report["start_date"],
            "end_date": first_report["end_date"],
            "employee_count": len(employee_reports),
            "employee_ids": employee_ids,
            "employee_reports": employee_reports,
            "generated_at": datetime.now().isoformat()
        }

        return comparison_data

    def save_report(
        self,
        report_data: Dict[str, Any],
        is_comparison: bool = False
    ) -> int:
        """
        Save a report to the database.

        Args:
            report_data: Report data dictionary
            is_comparison: Whether this is a comparison report

        Returns:
            ID of the saved report

        Requirements: 13.2, 13.3
        """
        # Extract metadata
        if is_comparison:
            employee_id = None
            start_date_str = report_data["start_date"]
            end_date_str = report_data["end_date"]
            report_type_str = report_data["report_type"]
        else:
            employee_id = report_data["employee_id"]
            start_date_str = report_data["start_date"]
            end_date_str = report_data["end_date"]
            report_type_str = report_data["report_type"]

        # Parse dates
        start_date = date.fromisoformat(start_date_str)
        end_date = date.fromisoformat(end_date_str)

        # Parse report type
        report_type = ReportType(report_type_str)

        # Serialize report data to JSON
        data_json = json.dumps(report_data)

        # Create report record
        report = Report(
            employee_id=employee_id,
            report_type=report_type,
            start_date=start_date,
            end_date=end_date,
            data=data_json
        )

        self.db.add(report)
        self.db.commit()
        self.db.refresh(report)

        logger.info(f"Saved report with ID {report.id}")
        return report.id

    def load_report(self, report_id: int) -> Dict[str, Any]:
        """
        Load a report from the database.

        Args:
            report_id: ID of the report to load

        Returns:
            Report data dictionary

        Requirements: 15.1, 15.2
        """
        report = self.db.query(Report).filter(
            Report.id == report_id
        ).first()

        if not report:
            raise ValueError(f"Report with ID {report_id} not found")

        # Deserialize JSON data
        report_data = json.loads(report.data)

        # Add metadata
        report_data["report_id"] = report.id
        report_data["created_at"] = report.created_at.isoformat()

        return report_data

    def list_reports(
        self,
        employee_id: Optional[int] = None,
        report_type: Optional[ReportType] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        limit: Optional[int] = None
    ) -> List[Dict[str, Any]]:
        """
        List reports with optional filters.

        Args:
            employee_id: Filter by employee ID
            report_type: Filter by report type
            start_date: Filter by reports created after this date
            end_date: Filter by reports created before this date
            limit: Maximum number of reports to return

        Returns:
            List of report metadata dictionaries

        Requirements: 13.5, 18.1
        """
        query = self.db.query(Report)

        # Apply filters
        if employee_id is not None:
            query = query.filter(Report.employee_id == employee_id)

        if report_type is not None:
            query = query.filter(Report.report_type == report_type)

        if start_date is not None:
            query = query.filter(Report.created_at >= start_date)

        if end_date is not None:
            query = query.filter(Report.created_at <= end_date)

        # Sort by creation date (newest first)
        query = query.order_by(Report.created_at.desc())

        # Apply limit
        if limit is not None:
            query = query.limit(limit)

        # Execute query
        reports = query.all()

        # Build metadata list
        report_list = []
        for report in reports:
            # Parse data to get employee name
            report_data = json.loads(report.data)

            metadata = {
                "report_id": report.id,
                "employee_id": report.employee_id,
                "employee_name": report_data.get("employee_name"),
                "report_type": report.report_type.value,
                "start_date": report.start_date.isoformat(),
                "end_date": report.end_date.isoformat(),
                "created_at": report.created_at.isoformat(),
                "is_comparison": report.employee_id is None
            }
            report_list.append(metadata)

        return report_list

    def export_report_json(self, report_data: Dict[str, Any]) -> str:
        """
        Export a report to JSON format.

        Args:
            report_data: Report data dictionary

        Returns:
            JSON string

        Requirements: 14.1, 14.4
        """
        # Create a clean copy for export
        export_data = {
            "report_metadata": {
                "employee_id": report_data.get("employee_id"),
                "employee_name": report_data.get("employee_name"),
                "position": report_data.get("position"),
                "report_type": report_data.get("report_type"),
                "start_date": report_data.get("start_date"),
                "end_date": report_data.get("end_date"),
                "generated_at": report_data.get("generated_at"),
                "report_id": report_data.get("report_id"),
                "created_at": report_data.get("created_at")
            },
            "quotas": report_data.get("quotas", {}),
            "ratio_descriptions": report_data.get("ratio_descriptions", {}),
            "aggregated_data": report_data.get("aggregated_data", {}),
            "employee_reports": report_data.get("employee_reports", [])
        }

        # Serialize to JSON with pretty printing
        return json.dumps(export_data, indent=2, ensure_ascii=False)

    def export_report_excel(self, report_data: Dict[str, Any]) -> bytes:
        """
        Export a report to Excel format.

        Args:
            report_data: Report data dictionary

        Returns:
            Excel file as bytes

        Requirements: 14.1, 14.3
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install it with: pip install openpyxl"
            )

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        # Define styles
        header_font = Font(bold=True, size=14)
        section_font = Font(bold=True, size=12)
        header_fill = PatternFill(
            start_color="366092",
            end_color="366092",
            fill_type="solid"
        )
        header_font_white = Font(bold=True, color="FFFFFF")

        # Add report metadata
        row = 1
        ws[f"A{row}"] = "Mitarbeiter / Team Auswertung"
        ws[f"A{row}"].font = header_font
        row += 2

        # Metadata section
        metadata = [
            ("Mitarbeiter", report_data.get("employee_name", "N/A")),
            ("Position", report_data.get("position", "N/A")),
            ("Berichtstyp", report_data.get("report_type", "N/A")),
            (
                "Zeitraum",
                f"{report_data.get('start_date')} bis "
                f"{report_data.get('end_date')}"
            ),
            ("Erstellt am", report_data.get("generated_at", "N/A"))
        ]

        for label, value in metadata:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = value
            row += 1

        row += 2

        # Quotas section
        ws[f"A{row}"] = "Quoten"
        ws[f"A{row}"].font = section_font
        row += 1

        # Quota headers
        ws[f"A{row}"] = "Status"
        ws[f"B{row}"] = "Quote n %"
        ws[f"C{row}"] = "Verhältnis"
        for col in ["A", "B", "C"]:
            ws[f"{col}{row}"].font = header_font_white
            ws[f"{col}{row}"].fill = header_fill
        row += 1

        # Quota data
        quotas = report_data.get("quotas", {})
        ratios = report_data.get("ratio_descriptions", {})
        for quota_name, quota_value in quotas.items():
            ws[f"A{row}"] = quota_name
            ws[f"B{row}"] = f"{quota_value:.2f}%"
            ws[f"C{row}"] = ratios.get(quota_name, "")
            row += 1

        row += 2

        # Raw data section
        aggregated_data = report_data.get("aggregated_data", {})
        raw_data = aggregated_data.get("raw_data", {})

        if raw_data:
            ws[f"A{row}"] = "Leistungsdaten"
            ws[f"A{row}"].font = section_font
            row += 1

            # Raw data headers
            ws[f"A{row}"] = "Criterion"
            ws[f"B{row}"] = "Value"
            for col in ["A", "B"]:
                ws[f"{col}{row}"].font = header_font_white
                ws[f"{col}{row}"].fill = header_fill
            row += 1

            # Raw data values
            for criterion_name, value in raw_data.items():
                ws[f"A{row}"] = criterion_name
                ws[f"B{row}"] = value
                row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 50

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def export_report_pdf(self, report_data: Dict[str, Any]) -> bytes:
        """
        Export a report to PDF format.

        Args:
            report_data: Report data dictionary

        Returns:
            PDF file as bytes

        Requirements: 14.1, 14.2
        """
        if not REPORTLAB_AVAILABLE:
            raise ImportError(
                "reportlab is required for PDF export. "
                "Install it with: pip install reportlab"
            )

        # Lade Farbschema
        color_scheme = get_color_scheme()

        # Create PDF buffer
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=1*cm,
            leftMargin=1*cm,
            topMargin=1*cm,
            bottomMargin=1*cm
        )

        # Container for PDF elements
        elements = []
        styles = getSampleStyleSheet()

        # Custom styles mit Farbschema
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor(color_scheme.title_color),
            spaceAfter=30
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor(color_scheme.primary_color),
            spaceAfter=12
        )

        # Title
        elements.append(Paragraph("Mitarbeiter / Team Auswertung", title_style))
        elements.append(Spacer(1, 0.5*cm))

        # Metadata
        agent_name_line = ""
        if report_data.get('agent_name'):
            agent_name_line = f"<b>Agentenname:</b> {report_data.get('agent_name')}<br/>"
        
        # Formatiere Daten
        report_type_name = get_report_type_name_german(report_data.get('report_type', ''))
        
        # Parse und formatiere Datumsstrings
        try:
            start_date_obj = date.fromisoformat(report_data.get('start_date'))
            start_date_formatted = format_date_german(start_date_obj)
        except:
            start_date_formatted = report_data.get('start_date', 'N/A')
        
        try:
            end_date_obj = date.fromisoformat(report_data.get('end_date'))
            end_date_formatted = format_date_german(end_date_obj)
        except:
            end_date_formatted = report_data.get('end_date', 'N/A')
        
        created_at_formatted = format_datetime_german(report_data.get('generated_at', ''))
        
        # Team information (if available)
        team_line = ""
        if report_data.get('team_name'):
            team_line = f"<b>Team:</b> {report_data.get('team_name')}<br/>"
            if report_data.get('team_leader'):
                team_line += f"<b>Teamleiter:</b> {report_data.get('team_leader')}<br/>"
        
        metadata_text = f"""
        <b>Mitarbeiter:</b> {report_data.get('employee_name', 'N/A')}<br/>
        {agent_name_line}
        <b>Position:</b> {report_data.get('position', 'N/A')}<br/>
        {team_line}
        <b>Berichtstyp:</b> {report_type_name}<br/>
        <b>Zeitraum:</b> {start_date_formatted} - {end_date_formatted}<br/>
        <b>Erstellt am:</b> {created_at_formatted}
        """
        elements.append(Paragraph(metadata_text, styles['Normal']))
        elements.append(Spacer(1, 1*cm))

        # Quotas section (mit KeepTogether für Überschrift + Tabelle)
        quotas_section = []
        quotas_section.append(Paragraph("Leistungsquoten", heading_style))

        quotas = report_data.get("quotas", {})
        ratios = report_data.get("ratio_descriptions", {})

        if quotas:
            # Create quota table
            quota_data = [["Status", "Quote in %", "Verhältnis"]]
            for quota_name, quota_value in quotas.items():
                quota_data.append([
                    quota_name,
                    f"{quota_value:.2f}%",
                    ratios.get(quota_name, "")
                ])

            quota_table = Table(quota_data, colWidths=[8*cm, 4*cm, 7*cm])
            quota_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(color_scheme.table_header_bg)),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor(color_scheme.table_row_bg)),
                ('FONTNAME', (1, 1), (1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (2, 1), (2, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor(color_scheme.border_color)),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            quotas_section.append(quota_table)
        
        # KeepTogether verhindert Seitenumbruch zwischen Überschrift und Tabelle
        elements.append(KeepTogether(quotas_section))
        elements.append(Spacer(1, 1*cm))

        # Raw data section (mit KeepTogether)
        aggregated_data = report_data.get("aggregated_data", {})
        raw_data = aggregated_data.get("raw_data", {})

        if raw_data:
            raw_data_section = []
            raw_data_section.append(Paragraph("Leistungsdaten", heading_style))

            # Create raw data table
            raw_data_table = [["Bezeichnung", "Anzahl"]]
            for criterion_name, value in raw_data.items():
                # Format value: integer + " Stück"
                try:
                    formatted_value = f"{int(float(value))} Stück"
                except (ValueError, TypeError):
                    formatted_value = f"{value} Stück"
                raw_data_table.append([criterion_name, formatted_value])

            data_table = Table(raw_data_table, colWidths=[13*cm, 6*cm])
            data_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(color_scheme.table_header_bg)),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor(color_scheme.table_row_bg)),
                ('FONTNAME', (1, 1), (1, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor(color_scheme.border_color)),
            ]))
            raw_data_section.append(data_table)
            
            # KeepTogether für Überschrift + Tabelle
            elements.append(KeepTogether(raw_data_section))
            elements.append(Spacer(1, 1*cm))

        # Add charts if available
        if PLOTLY_AVAILABLE:
            try:
                from controlling.chart_generator import ChartGenerator
                chart_gen = ChartGenerator()
                
                # Generate dashboard charts
                figures = chart_gen.create_dashboard(report_data)
                
                if figures:
                    # Create chart section with KeepTogether to prevent orphaned heading
                    chart_elements = []
                    chart_elements.append(Paragraph("Diagramme", heading_style))
                    chart_elements.append(Spacer(1, 0.5*cm))
                    
                    # Convert each plotly figure to image and add to section
                    for fig in figures:
                        try:
                            # Convert plotly figure to PNG bytes
                            img_bytes = pio.to_image(fig, format='png', width=800, height=500)
                            
                            # Create ReportLab Image from bytes
                            img_buffer = io.BytesIO(img_bytes)
                            img = Image(img_buffer, width=17*cm, height=10.5*cm)
                            
                            chart_elements.append(img)
                            chart_elements.append(Spacer(1, 0.5*cm))
                        except Exception as e:
                            logger.warning(f"Fehler beim Hinzufügen des Diagramms: {e}")
                    
                    # Add entire chart section wrapped in KeepTogether
                    elements.append(KeepTogether(chart_elements))
            except Exception as e:
                logger.warning(f"Fehler beim Generieren der Diagramme: {e}")

        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()
    
    def export_team_report_to_pdf(
        self,
        team_report_data: Dict[str, Any]
    ) -> bytes:
        """
        Exportiere Team-Bericht als PDF mit Bytes-Rückgabe.
        
        Args:
            team_report_data: Team-Berichtsdaten
            
        Returns:
            PDF als Bytes
        """
        if not REPORTLAB_AVAILABLE:
            raise ImportError(
                "reportlab is required for PDF export. "
                "Install it with: pip install reportlab"
            )
        
        # Lade Farbschema
        color_scheme = get_color_scheme()
        
        # Create PDF buffer
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        # Container for PDF elements
        elements = []
        styles = getSampleStyleSheet()
        
        # Custom styles mit Farbschema
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor(color_scheme.title_color),
            spaceAfter=30
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor(color_scheme.primary_color),
            spaceAfter=12
        )
        subheading_style = ParagraphStyle(
            'CustomSubHeading',
            parent=styles['Heading3'],
            fontSize=14,
            textColor=colors.HexColor(color_scheme.secondary_color),
            spaceAfter=10
        )
        
        # Title
        elements.append(Paragraph("Team-Auswertung", title_style))
        elements.append(Spacer(1, 0.5*cm))
        
        # Metadata
        # Formatiere Datumsstrings
        try:
            start_date_obj = date.fromisoformat(team_report_data.get('start_date'))
            start_date_formatted = format_date_german(start_date_obj)
        except:
            start_date_formatted = team_report_data.get('start_date', 'N/A')
        
        try:
            end_date_obj = date.fromisoformat(team_report_data.get('end_date'))
            end_date_formatted = format_date_german(end_date_obj)
        except:
            end_date_formatted = team_report_data.get('end_date', 'N/A')
        
        created_at_formatted = format_datetime_german(team_report_data.get('generated_at', ''))
        
        metadata_text = f"""
        <b>Team:</b> {team_report_data.get('team_name', 'N/A')}<br/>
        <b>Teamleiter:</b> {team_report_data.get('team_leader', 'Nicht zugewiesen')}<br/>
        <b>Position:</b> {team_report_data.get('position_name', 'N/A')}<br/>
        <b>Anzahl Mitarbeiter:</b> {team_report_data.get('employee_count', 0)}<br/>
        <b>Zeitraum:</b> {start_date_formatted} - {end_date_formatted}<br/>
        <b>Erstellt am:</b> {created_at_formatted}
        """
        elements.append(Paragraph(metadata_text, styles['Normal']))
        elements.append(Spacer(1, 1*cm))
        
        # Team-Quotas
        elements.append(Paragraph("Team-Leistungsquoten (Gesamt)", heading_style))
        
        team_quotas = team_report_data.get("team_quotas", {})
        team_ratios = team_report_data.get("team_ratio_descriptions", {})
        
        if team_quotas:
            quota_data = [["Quote", "Prozentsatz", "Verhältnis"]]
            for quota_name, quota_value in team_quotas.items():
                quota_data.append([
                    quota_name,
                    f"{quota_value:.2f}%",
                    team_ratios.get(quota_name, "")
                ])
            
            quota_table = Table(quota_data, colWidths=[7*cm, 3*cm, 7*cm])
            quota_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(color_scheme.table_header_bg)),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor(color_scheme.table_row_bg)),
                ('FONTSIZE', (2, 1), (2, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor(color_scheme.border_color)),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            elements.append(quota_table)
            elements.append(Spacer(1, 1*cm))
        
        # Statistiken
        statistics = team_report_data.get("statistics", {})
        quota_stats = statistics.get("quota_statistics", {})
        
        if quota_stats:
            elements.append(Paragraph("Statistiken & Leistungsvergleich", heading_style))
            
            stats_data = [["Quote", "Durchschnitt", "Min", "Max", "Bester", "Schlechtester"]]
            for quota_name, stats in quota_stats.items():
                stats_data.append([
                    quota_name,
                    f"{stats['average']:.2f}%",
                    f"{stats['min']:.2f}%",
                    f"{stats['max']:.2f}%",
                    stats['best_performer'],
                    stats['worst_performer']
                ])
            
            stats_table = Table(stats_data, colWidths=[4*cm, 2.5*cm, 2*cm, 2*cm, 3*cm, 3*cm])
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(color_scheme.table_header_bg)),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor(color_scheme.header_text_color)),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor(color_scheme.table_row_bg)),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor(color_scheme.border_color)),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            elements.append(stats_table)
            elements.append(Spacer(1, 1*cm))
        
        # Einzelne Mitarbeiter
        employees = team_report_data.get("employees", [])
        
        if employees:
            elements.append(Paragraph("Einzelne Mitarbeiter", heading_style))
            
            for i, emp in enumerate(employees):
                # Mitarbeiter-Name
                emp_name = f"{emp.get('name', 'N/A')}"
                if emp.get('agent_name'):
                    emp_name += f" ({emp.get('agent_name')})"
                
                elements.append(Paragraph(emp_name, subheading_style))
                
                # Mitarbeiter-Quotas
                emp_quotas = emp.get("quotas", {})
                emp_ratios = emp.get("ratio_descriptions", {})
                
                if emp_quotas:
                    emp_quota_data = [["Quote", "Prozentsatz"]]
                    for quota_name, quota_value in emp_quotas.items():
                        emp_quota_data.append([
                            quota_name,
                            f"{quota_value:.2f}%"
                        ])
                    
                    emp_table = Table(emp_quota_data, colWidths=[10*cm, 7*cm])
                    emp_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(color_scheme.secondary_color)),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor(color_scheme.table_alt_row_bg)),
                        ('FONTNAME', (1, 1), (1, -1), 'Helvetica-Bold'),
                        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor(color_scheme.grid_color)),
                        ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ]))
                    elements.append(emp_table)
                    elements.append(Spacer(1, 0.5*cm))
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()
    
    def export_comparison_report_to_pdf(
        self,
        comparison_data: Dict[str, Any]
    ) -> bytes:
        """
        Exportiere Mitarbeiter-Vergleichsbericht als PDF mit Bytes-Rückgabe.
        
        Args:
            comparison_data: Vergleichsdaten
            
        Returns:
            PDF als Bytes
        """
        if not REPORTLAB_AVAILABLE:
            raise ImportError(
                "reportlab is required for PDF export. "
                "Install it with: pip install reportlab"
            )
        
        # Lade Farbschema
        color_scheme = get_color_scheme()
        
        # Create PDF buffer
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=1*cm,
            leftMargin=1*cm,
            topMargin=1*cm,
            bottomMargin=1*cm
        )
        
        # Container for PDF elements
        elements = []
        styles = getSampleStyleSheet()
        
        # Custom styles mit Farbschema
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor(color_scheme.title_color),
            spaceAfter=30
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor(color_scheme.primary_color),
            spaceAfter=12
        )
        
        # Title
        elements.append(Paragraph("Mitarbeiter-Vergleich", title_style))
        elements.append(Spacer(1, 0.5*cm))
        
        # Metadata
        # Formatiere Datumsstrings
        try:
            start_date_obj = date.fromisoformat(comparison_data.get('start_date'))
            start_date_formatted = format_date_german(start_date_obj)
        except:
            start_date_formatted = comparison_data.get('start_date', 'N/A')
        
        try:
            end_date_obj = date.fromisoformat(comparison_data.get('end_date'))
            end_date_formatted = format_date_german(end_date_obj)
        except:
            end_date_formatted = comparison_data.get('end_date', 'N/A')
        
        created_at_formatted = format_datetime_german(comparison_data.get('generated_at', ''))
        
        metadata_text = f"""
        <b>Team:</b> {comparison_data.get('team_name', 'N/A')}<br/>
        <b>Teamleiter:</b> {comparison_data.get('team_leader', 'Nicht zugewiesen')}<br/>
        <b>Position:</b> {comparison_data.get('position_name', 'N/A')}<br/>
        <b>Anzahl Mitarbeiter:</b> {comparison_data.get('employee_count', 0)}<br/>
        <b>Verglichene Mitarbeiter:</b> {len(comparison_data.get('employees', []))}<br/>
        <b>Zeitraum:</b> {start_date_formatted} - {end_date_formatted}<br/>
        <b>Erstellt am:</b> {created_at_formatted}
        """
        elements.append(Paragraph(metadata_text, styles['Normal']))
        elements.append(Spacer(1, 1*cm))
        
        # Rankings für jede Quote
        comparison_stats = comparison_data.get("comparison_statistics", {})
        rankings = comparison_stats.get("rankings", {})
        
        if rankings:
            elements.append(Paragraph("Leistungsranking", heading_style))
            
            for quota_name, ranking_list in rankings.items():
                elements.append(Paragraph(quota_name, styles['Heading3']))
                
                ranking_data = [["Rang", "Mitarbeiter", "Prozentsatz"]]
                for item in ranking_list:
                    ranking_data.append([
                        str(item['rank']),
                        item['name'],
                        f"{item['value']:.2f}%"
                    ])
                
                ranking_table = Table(ranking_data, colWidths=[2*cm, 10*cm, 5*cm])
                ranking_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(color_scheme.table_header_bg)),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                    ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                    ('ALIGN', (1, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 11),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor(color_scheme.table_row_bg)),
                    ('FONTNAME', (2, 1), (2, -1), 'Helvetica-Bold'),
                    ('GRID', (0, 0), (-1, -1), 1, colors.HexColor(color_scheme.border_color)),
                ]))
                elements.append(ranking_table)
                elements.append(Spacer(1, 0.7*cm))
        
        # Unterschiede zwischen Bestem und Schlechtestem
        differences = comparison_stats.get("differences", {})
        
        if differences:
            elements.append(Paragraph("Leistungsunterschiede", heading_style))
            
            diff_data = [["Quote", "Bester", "Prozentsatz", "Schlechtester", "Prozentsatz", "Differenz"]]
            for quota_name, diff_info in differences.items():
                diff_data.append([
                    quota_name,
                    diff_info['leader'],
                    f"{diff_info['leader_value']:.2f}%",
                    diff_info['last'],
                    f"{diff_info['last_value']:.2f}%",
                    f"{diff_info['difference']:.2f}%"
                ])
            
            diff_table = Table(diff_data, colWidths=[4*cm, 3.5*cm, 2*cm, 3.5*cm, 2*cm, 2*cm])
            diff_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(color_scheme.table_header_bg)),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor(color_scheme.table_row_bg)),
                ('FONTNAME', (2, 1), (2, -1), 'Helvetica-Bold'),
                ('FONTNAME', (4, 1), (4, -1), 'Helvetica-Bold'),
                ('FONTNAME', (5, 1), (5, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor(color_scheme.border_color)),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            elements.append(diff_table)
            elements.append(Spacer(1, 1*cm))
        
        # Detaillierte Mitarbeiter-Daten
        employees = comparison_data.get("employees", [])
        
        if employees:
            elements.append(Paragraph("Detaillierte Mitarbeiter-Daten", heading_style))
            
            for emp in employees:
                emp_data_rows = [["Mitarbeiter", emp.get('name', 'N/A')]]
                emp_data_rows.append(["Position", emp.get('position', 'N/A')])
                
                # Quotas
                for quota_name, quota_value in emp.get('quotas', {}).items():
                    emp_data_rows.append([quota_name, f"{quota_value:.2f}%"])
                
                emp_detail_table = Table(emp_data_rows, colWidths=[10*cm, 7*cm])
                emp_detail_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(color_scheme.secondary_color)),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor(color_scheme.header_text_color)),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor(color_scheme.table_alt_row_bg)),
                    ('GRID', (0, 0), (-1, -1), 1, colors.HexColor(color_scheme.grid_color)),
                ]))
                elements.append(emp_detail_table)
                elements.append(Spacer(1, 0.5*cm))
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()
    
    def export_ranking_report_to_pdf(
        self,
        ranking_data: Dict[str, Any]
    ) -> bytes:
        """
        Exportiere Mitarbeiter-Ranking als PDF.
        
        Args:
            ranking_data: Ranking-Daten vom RankingSystem
            
        Returns:
            PDF als Bytes
        """
        if not REPORTLAB_AVAILABLE:
            raise ImportError(
                "reportlab is required for PDF export. "
                "Install it with: pip install reportlab"
            )
        
        # Lade Farbschema
        color_scheme = get_color_scheme()
        
        # Create PDF buffer
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=1*cm,
            leftMargin=1*cm,
            topMargin=1*cm,
            bottomMargin=1*cm
        )
        
        # Container for PDF elements
        elements = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor(color_scheme.title_color),
            spaceAfter=30
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor(color_scheme.primary_color),
            spaceAfter=12
        )
        subheading_style = ParagraphStyle(
            'CustomSubHeading',
            parent=styles['Heading3'],
            fontSize=14,
            textColor=colors.HexColor(color_scheme.secondary_color),
            spaceAfter=10
        )
        
        # Title
        elements.append(Paragraph("Mitarbeiter-Rangliste", title_style))
        elements.append(Spacer(1, 0.5*cm))
        
        # Metadata
        period_info = ""
        if ranking_data.get("period_name"):
            period_info = f"<b>Periode:</b> {ranking_data.get('period_name')}<br/>"
        
        # Formatiere Datumsstrings
        try:
            start_date_obj = date.fromisoformat(ranking_data.get('start_date'))
            start_date_formatted = format_date_german(start_date_obj)
        except:
            start_date_formatted = ranking_data.get('start_date', 'N/A')
        
        try:
            end_date_obj = date.fromisoformat(ranking_data.get('end_date'))
            end_date_formatted = format_date_german(end_date_obj)
        except:
            end_date_formatted = ranking_data.get('end_date', 'N/A')
        
        created_at_formatted = format_datetime_german(ranking_data.get('generated_at', ''))
        
        metadata_text = f"""
        <b>Position:</b> {ranking_data.get('position_name', 'N/A')}<br/>
        {period_info}
        <b>Zeitraum:</b> {start_date_formatted} - {end_date_formatted}<br/>
        <b>Anzahl Mitarbeiter:</b> {ranking_data.get('employee_count', 0)}<br/>
        <b>Erstellt am:</b> {created_at_formatted}
        """
        elements.append(Paragraph(metadata_text, styles['Normal']))
        elements.append(Spacer(1, 1*cm))
        
        # Gesamt-Ranking
        overall_ranking = ranking_data.get("overall_ranking", [])
        
        if overall_ranking:
            elements.append(Paragraph("Gesamt-Rangliste (Durchschnitt)", heading_style))
            
            overall_data = [["Rang", "Mitarbeiter", "Agentenname", "Durchschnitt", "Anzahl Quotas"]]
            for entry in overall_ranking:
                agent_name = entry.get("agent_name", "-")
                overall_data.append([
                    str(entry["rank"]),
                    entry["name"],
                    agent_name,
                    f"{entry['average_score']:.2f}%",
                    str(entry["quota_count"])
                ])
            
            overall_table = Table(overall_data, colWidths=[2*cm, 5*cm, 4*cm, 4*cm, 4*cm])
            
            # Basis-Styles
            table_styles = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(color_scheme.table_header_bg)),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                ('ALIGN', (1, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor(color_scheme.table_row_bg)),
                ('FONTNAME', (3, 1), (3, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor(color_scheme.border_color)),
                # Highlight Top 3
                ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor("#FFD700")),  # Gold
            ]
            
            # Silber und Bronze nur wenn genug Teilnehmer
            if len(overall_ranking) > 1:
                table_styles.append(('BACKGROUND', (0, 2), (-1, 2), colors.HexColor("#C0C0C0")))  # Silber
            if len(overall_ranking) > 2:
                table_styles.append(('BACKGROUND', (0, 3), (-1, 3), colors.HexColor("#CD7F32")))  # Bronze
            
            overall_table.setStyle(TableStyle(table_styles))
            
            elements.append(overall_table)
            elements.append(Spacer(1, 1*cm))
        
        # Rankings pro Quota
        quota_rankings = ranking_data.get("quota_rankings", {})
        
        if quota_rankings:
            elements.append(Paragraph("Rankings nach einzelnen Leistungskriterien", heading_style))
            elements.append(Spacer(1, 0.5*cm))
            
            for quota_name, ranking_list in quota_rankings.items():
                elements.append(Paragraph(quota_name, subheading_style))
                
                quota_data = [["Rang", "Mitarbeiter", "Agentenname", "Wert"]]
                for entry in ranking_list:
                    agent_name = entry.get("agent_name", "-")
                    quota_data.append([
                        str(entry["rank"]),
                        entry["name"],
                        agent_name,
                        f"{entry['value']:.2f}%"
                    ])
                
                quota_table = Table(quota_data, colWidths=[2*cm, 6*cm, 6*cm, 5*cm])
                quota_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(color_scheme.secondary_color)),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                    ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                    ('ALIGN', (1, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 11),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor(color_scheme.table_alt_row_bg)),
                    ('FONTNAME', (3, 1), (3, -1), 'Helvetica-Bold'),
                    ('GRID', (0, 0), (-1, -1), 1, colors.HexColor(color_scheme.grid_color)),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    # Highlight Platz 1
                    ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor("#FFFFCC")),
                ]))
                elements.append(quota_table)
                elements.append(Spacer(1, 0.7*cm))
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()
