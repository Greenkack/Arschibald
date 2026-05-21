# backend/services/report_generation_service.py

from typing import Dict, Any, List, Optional
from datetime import datetime
import uuid
import json
from pathlib import Path

from ..models.report_schemas import (
    ReportType, ReportFormat, ReportGenerationRequest,
    DetailedReportData, ExecutiveSummaryData, TechnicalReportData,
    FinancialReportData, EnvironmentalReportData,
    ReportMetadata, ReportResponse, ReportSection
)


class ReportGenerationService:
    """Service for generating various types of reports"""
    
    def __init__(self, output_dir: str = "reports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    async def generate_report(
        self,
        request: ReportGenerationRequest,
        project_data: Dict[str, Any],
        user_id: str
    ) -> ReportResponse:
        """Generate a report based on request"""
        
        # Generate unique report ID
        report_id = str(uuid.uuid4())
        
        # Prepare report data based on type
        if request.report_type == ReportType.DETAILED:
            report_data = self._prepare_detailed_report(project_data)
        elif request.report_type == ReportType.EXECUTIVE:
            report_data = self._prepare_executive_summary(project_data)
        elif request.report_type == ReportType.TECHNICAL:
            report_data = self._prepare_technical_report(project_data)
        elif request.report_type == ReportType.FINANCIAL:
            report_data = self._prepare_financial_report(project_data)
        elif request.report_type == ReportType.ENVIRONMENTAL:
            report_data = self._prepare_environmental_report(project_data)
        else:  # CUSTOM
            report_data = self._prepare_custom_report(project_data, request.custom_sections)
        
        # Generate report in requested format
        if request.format == ReportFormat.PDF:
            file_path = await self._generate_pdf_report(report_id, report_data, request)
        elif request.format == ReportFormat.HTML:
            file_path = await self._generate_html_report(report_id, report_data, request)
        elif request.format == ReportFormat.JSON:
            file_path = await self._generate_json_report(report_id, report_data, request)
        elif request.format == ReportFormat.EXCEL:
            file_path = await self._generate_excel_report(report_id, report_data, request)
        else:  # CSV
            file_path = await self._generate_csv_report(report_id, report_data, request)
        
        # Create metadata
        metadata = ReportMetadata(
            report_id=report_id,
            project_id=request.project_id,
            report_type=request.report_type,
            format=request.format,
            generated_at=datetime.now(),
            generated_by=user_id,
            file_size=file_path.stat().st_size if file_path.exists() else None
        )
        
        # Create response
        return ReportResponse(
            success=True,
            report_id=report_id,
            metadata=metadata,
            download_url=f"/api/v1/reports/{report_id}/download",
            preview_url=f"/api/v1/reports/{report_id}/preview",
            file_path=str(file_path),
            message="Report generated successfully"
        )
    
    def _prepare_detailed_report(self, project_data: Dict[str, Any]) -> DetailedReportData:
        """Prepare detailed report data"""
        return DetailedReportData(
            project_info={
                "name": project_data.get("name", ""),
                "customer": project_data.get("customer_name", ""),
                "location": project_data.get("location", ""),
                "date": datetime.now().isoformat()
            },
            system_configuration={
                "system_size": project_data.get("system_size", 0),
                "module_count": project_data.get("module_count", 0),
                "module_type": project_data.get("module_type", ""),
                "inverter": project_data.get("inverter", ""),
                "battery": project_data.get("battery", ""),
                "mounting_type": project_data.get("mounting_type", "")
            },
            calculation_results={
                "annual_production": project_data.get("annual_production", 0),
                "self_consumption": project_data.get("self_consumption_rate", 0),
                "grid_feed_in": project_data.get("grid_feed_in", 0),
                "performance_ratio": project_data.get("performance_ratio", 0)
            },
            energy_analysis={
                "monthly_production": project_data.get("monthly_production", []),
                "hourly_profile": project_data.get("hourly_profile", []),
                "seasonal_variation": project_data.get("seasonal_variation", {})
            },
            financial_analysis={
                "total_cost": project_data.get("total_cost", 0),
                "annual_savings": project_data.get("annual_savings", 0),
                "payback_period": project_data.get("payback_period", 0),
                "roi_25_years": project_data.get("roi_25_years", 0),
                "npv": project_data.get("npv", 0),
                "irr": project_data.get("irr", 0)
            },
            environmental_impact={
                "co2_savings_annual": project_data.get("co2_savings", 0),
                "co2_savings_25_years": project_data.get("co2_savings_25_years", 0),
                "trees_equivalent": project_data.get("trees_equivalent", 0),
                "cars_equivalent": project_data.get("cars_equivalent", 0)
            },
            technical_specifications={
                "roof_area": project_data.get("roof_area", 0),
                "roof_angle": project_data.get("roof_angle", 0),
                "orientation": project_data.get("orientation", ""),
                "shading_factor": project_data.get("shading_factor", 0)
            },
            recommendations=project_data.get("recommendations", []),
            charts=self._generate_detailed_charts(project_data),
            tables=self._generate_detailed_tables(project_data)
        )
    
    def _prepare_executive_summary(self, project_data: Dict[str, Any]) -> ExecutiveSummaryData:
        """Prepare executive summary data"""
        return ExecutiveSummaryData(
            project_name=project_data.get("name", ""),
            customer_name=project_data.get("customer_name", ""),
            system_size=project_data.get("system_size", 0),
            total_cost=project_data.get("total_cost", 0),
            annual_savings=project_data.get("annual_savings", 0),
            payback_period=project_data.get("payback_period", 0),
            roi_percentage=project_data.get("roi_percentage", 0),
            co2_reduction=project_data.get("co2_savings", 0),
            key_highlights=[
                f"System Size: {project_data.get('system_size', 0):.2f} kWp",
                f"Annual Production: {project_data.get('annual_production', 0):,.0f} kWh",
                f"Annual Savings: €{project_data.get('annual_savings', 0):,.2f}",
                f"Payback Period: {project_data.get('payback_period', 0):.1f} years",
                f"CO₂ Reduction: {project_data.get('co2_savings', 0):,.0f} kg/year"
            ],
            recommendation=self._generate_recommendation(project_data),
            charts=self._generate_executive_charts(project_data)
        )
    
    def _prepare_technical_report(self, project_data: Dict[str, Any]) -> TechnicalReportData:
        """Prepare technical report data"""
        return TechnicalReportData(
            system_design={
                "system_size": project_data.get("system_size", 0),
                "module_count": project_data.get("module_count", 0),
                "string_configuration": project_data.get("string_configuration", ""),
                "array_layout": project_data.get("array_layout", {})
            },
            component_specifications=[
                {
                    "type": "PV Module",
                    "model": project_data.get("module_type", ""),
                    "power": project_data.get("module_power", 0),
                    "efficiency": project_data.get("module_efficiency", 0),
                    "dimensions": project_data.get("module_dimensions", "")
                },
                {
                    "type": "Inverter",
                    "model": project_data.get("inverter", ""),
                    "power": project_data.get("inverter_power", 0),
                    "efficiency": project_data.get("inverter_efficiency", 0)
                }
            ],
            installation_requirements={
                "roof_type": project_data.get("roof_type", ""),
                "mounting_system": project_data.get("mounting_type", ""),
                "cable_length": project_data.get("cable_length", 0),
                "grounding": project_data.get("grounding_requirements", "")
            },
            electrical_design={
                "dc_voltage": project_data.get("dc_voltage", 0),
                "dc_current": project_data.get("dc_current", 0),
                "ac_voltage": project_data.get("ac_voltage", 0),
                "ac_current": project_data.get("ac_current", 0)
            },
            mounting_system={
                "type": project_data.get("mounting_type", ""),
                "rail_length": project_data.get("rail_length", 0),
                "clamps": project_data.get("clamp_count", 0),
                "hooks": project_data.get("hook_count", 0)
            },
            performance_calculations={
                "annual_yield": project_data.get("annual_production", 0),
                "specific_yield": project_data.get("specific_yield", 0),
                "performance_ratio": project_data.get("performance_ratio", 0),
                "losses": project_data.get("system_losses", {})
            },
            compliance_standards=[
                "VDE 0100",
                "VDE 0126-1-1",
                "EN 61215",
                "EN 61730"
            ],
            technical_drawings=[]
        )
    
    def _prepare_financial_report(self, project_data: Dict[str, Any]) -> FinancialReportData:
        """Prepare financial report data"""
        return FinancialReportData(
            investment_summary={
                "total_investment": project_data.get("total_cost", 0),
                "equipment_cost": project_data.get("equipment_cost", 0),
                "installation_cost": project_data.get("installation_cost", 0),
                "other_costs": project_data.get("other_costs", 0)
            },
            cost_breakdown={
                "modules": project_data.get("module_cost", 0),
                "inverter": project_data.get("inverter_cost", 0),
                "battery": project_data.get("battery_cost", 0),
                "mounting": project_data.get("mounting_cost", 0),
                "installation": project_data.get("installation_cost", 0),
                "permits": project_data.get("permit_cost", 0)
            },
            revenue_projections={
                "year_1": project_data.get("revenue_year_1", 0),
                "year_5": project_data.get("revenue_year_5", 0),
                "year_10": project_data.get("revenue_year_10", 0),
                "year_25": project_data.get("revenue_year_25", 0),
                "total_25_years": project_data.get("total_revenue_25_years", 0)
            },
            cash_flow_analysis={
                "annual_cash_flows": project_data.get("annual_cash_flows", []),
                "cumulative_cash_flow": project_data.get("cumulative_cash_flow", [])
            },
            roi_analysis={
                "payback_period": project_data.get("payback_period", 0),
                "roi_percentage": project_data.get("roi_percentage", 0),
                "npv": project_data.get("npv", 0),
                "irr": project_data.get("irr", 0)
            },
            financing_options=[
                {
                    "type": "Cash Purchase",
                    "down_payment": project_data.get("total_cost", 0),
                    "monthly_payment": 0,
                    "total_cost": project_data.get("total_cost", 0)
                },
                {
                    "type": "Loan (10 years, 3%)",
                    "down_payment": project_data.get("total_cost", 0) * 0.2,
                    "monthly_payment": self._calculate_loan_payment(
                        project_data.get("total_cost", 0) * 0.8, 0.03, 10
                    ),
                    "total_cost": project_data.get("total_cost", 0) * 1.15
                }
            ],
            tax_benefits={
                "depreciation": project_data.get("depreciation_benefit", 0),
                "investment_tax_credit": project_data.get("tax_credit", 0),
                "total_tax_savings": project_data.get("total_tax_savings", 0)
            },
            sensitivity_analysis={
                "electricity_price_increase": project_data.get("sensitivity_electricity", {}),
                "system_performance": project_data.get("sensitivity_performance", {}),
                "interest_rate": project_data.get("sensitivity_interest", {})
            },
            charts=self._generate_financial_charts(project_data)
        )
    
    def _prepare_environmental_report(self, project_data: Dict[str, Any]) -> EnvironmentalReportData:
        """Prepare environmental report data"""
        co2_annual = project_data.get("co2_savings", 0)
        
        return EnvironmentalReportData(
            co2_emissions_avoided=co2_annual,
            equivalent_trees_planted=int(co2_annual / 20),  # ~20kg CO2 per tree per year
            equivalent_cars_removed=int(co2_annual / 4600),  # ~4600kg CO2 per car per year
            renewable_energy_percentage=project_data.get("renewable_percentage", 0),
            lifecycle_analysis={
                "energy_payback_time": project_data.get("energy_payback_time", 0),
                "co2_payback_time": project_data.get("co2_payback_time", 0),
                "lifecycle_co2_savings": co2_annual * 25
            },
            environmental_certifications=[
                "ISO 14001",
                "Carbon Neutral Certified"
            ],
            sustainability_metrics={
                "renewable_energy_generated": project_data.get("annual_production", 0),
                "fossil_fuel_avoided": project_data.get("fossil_fuel_avoided", 0),
                "water_saved": project_data.get("water_saved", 0)
            },
            charts=self._generate_environmental_charts(project_data)
        )
    
    def _prepare_custom_report(
        self,
        project_data: Dict[str, Any],
        sections: Optional[List[ReportSection]]
    ) -> Dict[str, Any]:
        """Prepare custom report data"""
        if not sections:
            sections = []
        
        return {
            "project_data": project_data,
            "sections": [section.dict() for section in sections],
            "generated_at": datetime.now().isoformat()
        }
    
    def _generate_recommendation(self, project_data: Dict[str, Any]) -> str:
        """Generate recommendation based on project data"""
        roi = project_data.get("roi_percentage", 0)
        payback = project_data.get("payback_period", 0)
        
        if roi > 200 and payback < 8:
            return "Highly Recommended: Excellent financial returns with short payback period."
        elif roi > 150 and payback < 10:
            return "Recommended: Good financial returns with reasonable payback period."
        elif roi > 100 and payback < 12:
            return "Acceptable: Positive returns with moderate payback period."
        else:
            return "Consider: Returns are positive but payback period is longer."
    
    def _calculate_loan_payment(self, principal: float, annual_rate: float, years: int) -> float:
        """Calculate monthly loan payment"""
        monthly_rate = annual_rate / 12
        num_payments = years * 12
        
        if monthly_rate == 0:
            return principal / num_payments
        
        payment = principal * (monthly_rate * (1 + monthly_rate) ** num_payments) / \
                  ((1 + monthly_rate) ** num_payments - 1)
        return payment
    
    def _generate_detailed_charts(self, project_data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Generate charts for detailed report"""
        return [
            {
                "type": "line",
                "title": "Monthly Energy Production",
                "data": project_data.get("monthly_production", [])
            },
            {
                "type": "bar",
                "title": "Cost Breakdown",
                "data": project_data.get("cost_breakdown", {})
            },
            {
                "type": "area",
                "title": "Cumulative Savings Over 25 Years",
                "data": project_data.get("cumulative_savings", [])
            }
        ]
    
    def _generate_executive_charts(self, project_data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Generate charts for executive summary"""
        return [
            {
                "type": "pie",
                "title": "Cost Distribution",
                "data": project_data.get("cost_breakdown", {})
            },
            {
                "type": "bar",
                "title": "Annual Savings",
                "data": {"Savings": project_data.get("annual_savings", 0)}
            }
        ]
    
    def _generate_financial_charts(self, project_data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Generate charts for financial report"""
        return [
            {
                "type": "waterfall",
                "title": "Cash Flow Analysis",
                "data": project_data.get("annual_cash_flows", [])
            },
            {
                "type": "line",
                "title": "Cumulative Cash Flow",
                "data": project_data.get("cumulative_cash_flow", [])
            }
        ]
    
    def _generate_environmental_charts(self, project_data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Generate charts for environmental report"""
        return [
            {
                "type": "bar",
                "title": "CO₂ Savings Over Time",
                "data": project_data.get("co2_savings_timeline", [])
            },
            {
                "type": "pie",
                "title": "Environmental Impact Distribution",
                "data": project_data.get("environmental_impact", {})
            }
        ]
    
    def _generate_detailed_tables(self, project_data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Generate tables for detailed report"""
        return [
            {
                "title": "System Components",
                "headers": ["Component", "Model", "Quantity", "Unit Price", "Total"],
                "rows": project_data.get("component_list", [])
            },
            {
                "title": "Annual Performance",
                "headers": ["Year", "Production (kWh)", "Savings (€)", "Cumulative Savings (€)"],
                "rows": project_data.get("annual_performance", [])
            }
        ]
    
    async def _generate_pdf_report(
        self,
        report_id: str,
        report_data: Any,
        request: ReportGenerationRequest
    ) -> Path:
        """Generate PDF report"""
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import mm
        
        file_path = self.output_dir / f"{report_id}.pdf"
        
        # Create PDF
        pdf = canvas.Canvas(str(file_path), pagesize=A4)
        width, height = A4
        
        # Add title
        pdf.setFont("Helvetica-Bold", 24)
        pdf.drawString(30*mm, height - 30*mm, f"{request.report_type.value.title()} Report")
        
        # Add content based on report type
        y_position = height - 50*mm
        pdf.setFont("Helvetica", 12)
        
        if isinstance(report_data, dict):
            for key, value in report_data.items():
                if y_position < 30*mm:
                    pdf.showPage()
                    y_position = height - 30*mm
                
                pdf.drawString(30*mm, y_position, f"{key}: {value}")
                y_position -= 6*mm
        
        pdf.save()
        return file_path
    
    async def _generate_html_report(
        self,
        report_id: str,
        report_data: Any,
        request: ReportGenerationRequest
    ) -> Path:
        """Generate HTML report"""
        file_path = self.output_dir / f"{report_id}.html"
        
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>{request.report_type.value.title()} Report</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 40px; }}
                h1 {{ color: #333; }}
                .section {{ margin: 20px 0; }}
                table {{ border-collapse: collapse; width: 100%; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #4CAF50; color: white; }}
            </style>
        </head>
        <body>
            <h1>{request.report_type.value.title()} Report</h1>
            <div class="section">
                <pre>{json.dumps(report_data.dict() if hasattr(report_data, 'dict') else report_data, indent=2)}</pre>
            </div>
        </body>
        </html>
        """
        
        file_path.write_text(html_content)
        return file_path
    
    async def _generate_json_report(
        self,
        report_id: str,
        report_data: Any,
        request: ReportGenerationRequest
    ) -> Path:
        """Generate JSON report"""
        file_path = self.output_dir / f"{report_id}.json"
        
        data = report_data.dict() if hasattr(report_data, 'dict') else report_data
        file_path.write_text(json.dumps(data, indent=2, default=str))
        
        return file_path
    
    async def _generate_excel_report(
        self,
        report_id: str,
        report_data: Any,
        request: ReportGenerationRequest
    ) -> Path:
        """Generate Excel report"""
        import openpyxl
        from openpyxl.styles import Font, Alignment
        
        file_path = self.output_dir / f"{report_id}.xlsx"
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = request.report_type.value.title()
        
        # Add title
        ws['A1'] = f"{request.report_type.value.title()} Report"
        ws['A1'].font = Font(size=16, bold=True)
        
        # Add data
        row = 3
        data = report_data.dict() if hasattr(report_data, 'dict') else report_data
        for key, value in data.items():
            ws[f'A{row}'] = str(key)
            ws[f'B{row}'] = str(value)
            row += 1
        
        wb.save(file_path)
        return file_path
    
    async def _generate_csv_report(
        self,
        report_id: str,
        report_data: Any,
        request: ReportGenerationRequest
    ) -> Path:
        """Generate CSV report"""
        import csv
        
        file_path = self.output_dir / f"{report_id}.csv"
        
        data = report_data.dict() if hasattr(report_data, 'dict') else report_data
        
        with open(file_path, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(['Field', 'Value'])
            
            for key, value in data.items():
                writer.writerow([key, value])
        
        return file_path
