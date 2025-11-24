"""
Results export service for multiple formats.
"""

import io
import json
import csv
import uuid
from datetime import datetime, timedelta
from typing import Dict, Any, List, Optional, BinaryIO
from pathlib import Path
import xml.etree.ElementTree as ET
import xml.dom.minidom as minidom

from reportlab.lib.pagesizes import A4, letter, legal
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.pdfgen import canvas

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter

from ..models.export_schemas import (
    ExportRequest, ExportResponse, PDFExportOptions, ExcelExportOptions,
    CSVExportOptions, JSONExportOptions, XMLExportOptions, BatchExportRequest
)


class ExportService:
    """Service for exporting results in multiple formats"""
    
    def __init__(self, export_dir: str = "exports"):
        self.export_dir = Path(export_dir)
        self.export_dir.mkdir(exist_ok=True)
        self.exports_cache: Dict[str, Dict[str, Any]] = {}
    
    async def export_result(self, request: ExportRequest, result_data: Dict[str, Any]) -> ExportResponse:
        """Export result in specified format"""
        export_id = str(uuid.uuid4())
        
        # Route to appropriate export method
        if request.format == 'pdf':
            file_data = await self._export_pdf(result_data, PDFExportOptions(**request.options))
            file_ext = 'pdf'
        elif request.format == 'excel':
            file_data = await self._export_excel(result_data, ExcelExportOptions(**request.options))
            file_ext = 'xlsx'
        elif request.format == 'csv':
            file_data = await self._export_csv(result_data, CSVExportOptions(**request.options))
            file_ext = 'csv'
        elif request.format == 'json':
            file_data = await self._export_json(result_data, JSONExportOptions(**request.options))
            file_ext = 'json'
        elif request.format == 'xml':
            file_data = await self._export_xml(result_data, XMLExportOptions(**request.options))
            file_ext = 'xml'
        else:
            raise ValueError(f"Unsupported export format: {request.format}")
        
        # Generate file name
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f"result_{request.result_id}_{timestamp}.{file_ext}"
        file_path = self.export_dir / file_name
        
        # Save file
        with open(file_path, 'wb') as f:
            f.write(file_data)
        
        # Cache export info
        self.exports_cache[export_id] = {
            'file_path': str(file_path),
            'file_name': file_name,
            'created_at': datetime.now(),
            'expires_at': datetime.now() + timedelta(hours=24)
        }
        
        return ExportResponse(
            export_id=export_id,
            format=request.format,
            file_name=file_name,
            file_size=len(file_data),
            download_url=f"/api/v1/exports/{export_id}/download",
            expires_at=datetime.now() + timedelta(hours=24)
        )
    
    async def _export_pdf(self, data: Dict[str, Any], options: PDFExportOptions) -> bytes:
        """Export result as PDF"""
        buffer = io.BytesIO()
        
        # Set page size
        page_size = {
            'A4': A4,
            'Letter': letter,
            'Legal': legal
        }[options.page_size]
        
        # Create PDF document
        doc = SimpleDocTemplate(
            buffer,
            pagesize=page_size,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18
        )
        
        # Build content
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1976d2'),
            spaceAfter=30
        )
        story.append(Paragraph(data.get('title', 'Calculation Results'), title_style))
        story.append(Spacer(1, 12))
        
        # Summary section
        if options.include_summary and 'summary' in data:
            story.append(Paragraph('Executive Summary', styles['Heading2']))
            story.append(Spacer(1, 12))
            for key, value in data['summary'].items():
                text = f"<b>{key}:</b> {self._format_value(value)}"
                story.append(Paragraph(text, styles['Normal']))
            story.append(Spacer(1, 20))
        
        # Data tables
        if options.include_tables and 'tables' in data:
            for table_name, table_data in data['tables'].items():
                story.append(Paragraph(table_name, styles['Heading2']))
                story.append(Spacer(1, 12))
                
                # Create table
                table = self._create_pdf_table(table_data)
                story.append(table)
                story.append(Spacer(1, 20))
        
        # Charts (as images)
        if options.include_charts and 'charts' in data:
            for chart_name, chart_path in data['charts'].items():
                story.append(Paragraph(chart_name, styles['Heading2']))
                story.append(Spacer(1, 12))
                if Path(chart_path).exists():
                    img = Image(chart_path, width=6*inch, height=4*inch)
                    story.append(img)
                story.append(Spacer(1, 20))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    
    def _create_pdf_table(self, table_data: List[List[Any]]) -> Table:
        """Create formatted PDF table"""
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        return table
    
    async def _export_excel(self, data: Dict[str, Any], options: ExcelExportOptions) -> bytes:
        """Export result as Excel"""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Summary sheet
        ws_summary = wb.create_sheet("Summary")
        self._write_excel_summary(ws_summary, data.get('summary', {}))
        
        # Data sheets
        if 'tables' in data:
            for i, (table_name, table_data) in enumerate(data['tables'].items()):
                sheet_name = options.sheet_names[i] if options.sheet_names and i < len(options.sheet_names) else table_name
                ws = wb.create_sheet(sheet_name[:31])  # Excel sheet name limit
                self._write_excel_table(ws, table_data, options)
        
        # Charts
        if options.include_charts and 'chart_data' in data:
            self._add_excel_charts(wb, data['chart_data'])
        
        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def _write_excel_summary(self, ws, summary: Dict[str, Any]):
        """Write summary data to Excel sheet"""
        # Header
        ws['A1'] = 'Metric'
        ws['B1'] = 'Value'
        ws['A1'].font = Font(bold=True, size=14)
        ws['B1'].font = Font(bold=True, size=14)
        
        # Data
        row = 2
        for key, value in summary.items():
            ws[f'A{row}'] = key
            ws[f'B{row}'] = self._format_value(value)
            row += 1
        
        # Formatting
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def _write_excel_table(self, ws, table_data: List[List[Any]], options: ExcelExportOptions):
        """Write table data to Excel sheet"""
        # Write data
        for row_idx, row_data in enumerate(table_data, start=1):
            for col_idx, cell_value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = cell_value
                
                # Header formatting
                if row_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
        
        # Freeze panes
        if options.freeze_panes:
            ws.freeze_panes = 'A2'
        
        # Auto-filter
        if options.auto_filter:
            ws.auto_filter.ref = ws.dimensions
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _add_excel_charts(self, wb, chart_data: Dict[str, Any]):
        """Add charts to Excel workbook"""
        ws = wb.create_sheet("Charts")
        
        # Example: Line chart
        if 'line_chart' in chart_data:
            chart = LineChart()
            chart.title = chart_data['line_chart'].get('title', 'Chart')
            chart.style = 13
            chart.y_axis.title = chart_data['line_chart'].get('y_label', 'Value')
            chart.x_axis.title = chart_data['line_chart'].get('x_label', 'Category')
            
            # Add data (simplified - would need actual data range)
            ws.append(['Category', 'Value'])
            for item in chart_data['line_chart'].get('data', []):
                ws.append([item.get('x'), item.get('y')])
            
            data = Reference(ws, min_col=2, min_row=1, max_row=len(chart_data['line_chart'].get('data', [])) + 1)
            cats = Reference(ws, min_col=1, min_row=2, max_row=len(chart_data['line_chart'].get('data', [])) + 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            ws.add_chart(chart, "D2")
    
    async def _export_csv(self, data: Dict[str, Any], options: CSVExportOptions) -> bytes:
        """Export result as CSV"""
        buffer = io.StringIO()
        
        # Get first table or flatten data
        if 'tables' in data and data['tables']:
            table_data = list(data['tables'].values())[0]
        else:
            # Flatten summary data
            table_data = [['Metric', 'Value']]
            for key, value in data.get('summary', {}).items():
                table_data.append([key, self._format_value(value, german=True)])
        
        # Write CSV
        writer = csv.writer(
            buffer,
            delimiter=options.delimiter,
            quoting=csv.QUOTE_MINIMAL
        )
        
        for row in table_data:
            # Format numbers with German separators if needed
            formatted_row = []
            for cell in row:
                if isinstance(cell, (int, float)):
                    formatted_cell = self._format_number_german(cell, options)
                else:
                    formatted_cell = str(cell)
                formatted_row.append(formatted_cell)
            writer.writerow(formatted_row)
        
        # Convert to bytes
        csv_string = buffer.getvalue()
        return csv_string.encode(options.encoding)
    
    def _format_number_german(self, value: float, options: CSVExportOptions) -> str:
        """Format number with German separators"""
        # Format with 2 decimal places
        formatted = f"{value:.2f}"
        
        # Split into integer and decimal parts
        parts = formatted.split('.')
        integer_part = parts[0]
        decimal_part = parts[1] if len(parts) > 1 else '00'
        
        # Add thousands separator
        integer_with_sep = ''
        for i, digit in enumerate(reversed(integer_part)):
            if i > 0 and i % 3 == 0:
                integer_with_sep = options.thousands_separator + integer_with_sep
            integer_with_sep = digit + integer_with_sep
        
        # Combine with decimal separator
        return f"{integer_with_sep}{options.decimal_separator}{decimal_part}"
    
    async def _export_json(self, data: Dict[str, Any], options: JSONExportOptions) -> bytes:
        """Export result as JSON"""
        export_data = data.copy()
        
        # Add metadata
        if options.include_metadata:
            export_data['_metadata'] = {
                'exported_at': datetime.now().isoformat() if options.date_format == 'iso' else datetime.now().timestamp(),
                'format': 'json',
                'version': '1.0'
            }
        
        # Convert dates
        export_data = self._convert_dates(export_data, options.date_format)
        
        # Serialize
        json_string = json.dumps(
            export_data,
            indent=2 if options.pretty_print else None,
            ensure_ascii=False
        )
        
        return json_string.encode('utf-8')
    
    async def _export_xml(self, data: Dict[str, Any], options: XMLExportOptions) -> bytes:
        """Export result as XML"""
        root = ET.Element(options.root_element)
        
        # Add metadata
        root.set('exported_at', datetime.now().isoformat())
        root.set('format', 'xml')
        
        # Convert data to XML
        self._dict_to_xml(data, root)
        
        # Convert to string
        xml_string = ET.tostring(root, encoding='unicode')
        
        # Pretty print if requested
        if options.pretty_print:
            dom = minidom.parseString(xml_string)
            xml_string = dom.toprettyxml(indent="  ")
        
        return xml_string.encode('utf-8')
    
    def _dict_to_xml(self, data: Dict[str, Any], parent: ET.Element):
        """Convert dictionary to XML elements"""
        for key, value in data.items():
            # Sanitize key for XML
            safe_key = key.replace(' ', '_').replace('-', '_')
            
            if isinstance(value, dict):
                child = ET.SubElement(parent, safe_key)
                self._dict_to_xml(value, child)
            elif isinstance(value, list):
                for item in value:
                    child = ET.SubElement(parent, safe_key)
                    if isinstance(item, dict):
                        self._dict_to_xml(item, child)
                    else:
                        child.text = str(item)
            else:
                child = ET.SubElement(parent, safe_key)
                child.text = str(value)
    
    def _format_value(self, value: Any, german: bool = False) -> str:
        """Format value for display"""
        if isinstance(value, float):
            if german:
                return self._format_number_german(value, CSVExportOptions())
            return f"{value:,.2f}"
        elif isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        return str(value)
    
    def _convert_dates(self, data: Any, date_format: str) -> Any:
        """Convert datetime objects in data structure"""
        if isinstance(data, dict):
            return {k: self._convert_dates(v, date_format) for k, v in data.items()}
        elif isinstance(data, list):
            return [self._convert_dates(item, date_format) for item in data]
        elif isinstance(data, datetime):
            if date_format == 'iso':
                return data.isoformat()
            elif date_format == 'unix':
                return data.timestamp()
            else:
                return data.strftime(date_format)
        return data
    
    async def batch_export(self, request: BatchExportRequest, results_data: List[Dict[str, Any]]) -> List[ExportResponse]:
        """Export multiple results"""
        exports = []
        
        for result_id, result_data in zip(request.result_ids, results_data):
            export_request = ExportRequest(
                result_id=result_id,
                format=request.format,
                options=request.options
            )
            export_response = await self.export_result(export_request, result_data)
            exports.append(export_response)
        
        return exports
    
    def get_export_file(self, export_id: str) -> Optional[Path]:
        """Get export file path"""
        if export_id in self.exports_cache:
            export_info = self.exports_cache[export_id]
            if datetime.now() < export_info['expires_at']:
                return Path(export_info['file_path'])
        return None
    
    def cleanup_expired_exports(self):
        """Remove expired export files"""
        now = datetime.now()
        expired_ids = [
            export_id for export_id, info in self.exports_cache.items()
            if now >= info['expires_at']
        ]
        
        for export_id in expired_ids:
            export_info = self.exports_cache[export_id]
            file_path = Path(export_info['file_path'])
            if file_path.exists():
                file_path.unlink()
            del self.exports_cache[export_id]
