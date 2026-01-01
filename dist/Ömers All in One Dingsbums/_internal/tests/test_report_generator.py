"""
Unit Tests for Report Generator

Tests report generation, saving, loading, and listing functionality
for the Employee Controlling System.

Requirements: 9.1, 13.2, 15.1
"""

import sys
from pathlib import Path
from datetime import date
import pytest
import json
import io

# Add project root to path
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from controlling.report_generator import ReportGenerator  # noqa: E402
from controlling.managers import (  # noqa: E402
    EmployeeManager,
    PositionManager,
    CriterionManager,
    PerformanceDataManager
)
from controlling.models import ReportType  # noqa: E402
from backend.core.database import SessionLocal, engine, Base  # noqa: E402


@pytest.fixture(scope="module")
def setup_database():
    """Setup database for all tests"""
    Base.metadata.create_all(bind=engine)
    yield
    Base.metadata.drop_all(bind=engine)


@pytest.fixture
def db_session(setup_database):
    """Create a database session for each test"""
    session = SessionLocal()
    yield session
    session.rollback()
    session.close()


@pytest.fixture
def report_generator(db_session):
    """Create a report generator instance"""
    return ReportGenerator(db_session)


@pytest.fixture
def sample_employee(db_session):
    """Create a sample employee with performance data"""
    # Create position (or get existing)
    pos_mgr = PositionManager(db_session)
    from controlling.models import Position  # noqa: E402
    position = db_session.query(Position).filter(
        Position.name == "Sales Representative"
    ).first()

    if not position:
        position = pos_mgr.create_position(
            "Sales Representative",
            "Sales position"
        )

    # Create employee
    emp_mgr = EmployeeManager(db_session)
    employee = emp_mgr.create_employee(
        first_name="John",
        last_name="Doe",
        city="Berlin",
        birth_date=date(1990, 1, 1),
        position_id=position.id,
        start_date=date(2023, 1, 1)
    )

    # Create criteria (or get existing)
    crit_mgr = CriterionManager(db_session)
    from controlling.models import Criterion  # noqa: E402
    criteria = {}
    for name in ["Verkauf", "Angefahrene Termine gesamt"]:
        criterion = db_session.query(Criterion).filter(
            Criterion.name == name
        ).first()

        if not criterion:
            criterion = crit_mgr.create_criterion(
                name=name,
                description=f"{name} criterion"
            )
        criteria[name] = criterion

    # Assign criteria to position
    pos_mgr.assign_criteria(
        position.id,
        [c.id for c in criteria.values()]
    )

    # Record performance data
    perf_mgr = PerformanceDataManager(db_session)
    today = date.today()
    perf_mgr.record_performance(
        employee.id,
        criteria["Verkauf"].id,
        10.0,
        today
    )
    perf_mgr.record_performance(
        employee.id,
        criteria["Angefahrene Termine gesamt"].id,
        40.0,
        today
    )

    return employee


class TestReportGeneration:
    """Test report generation for different time periods"""

    def test_generate_daily_report(self, report_generator, sample_employee):
        """Test daily report generation"""
        report = report_generator.generate_report(
            sample_employee.id,
            ReportType.DAILY
        )

        assert report["employee_id"] == sample_employee.id
        assert report["employee_name"] == sample_employee.full_name
        assert report["report_type"] == ReportType.DAILY.value
        assert "aggregated_data" in report
        assert "quotas" in report
        assert "ratio_descriptions" in report

    def test_generate_weekly_report(self, report_generator, sample_employee):
        """Test weekly report generation"""
        report = report_generator.generate_report(
            sample_employee.id,
            ReportType.WEEKLY
        )

        assert report["report_type"] == ReportType.WEEKLY.value
        # Verify date range is a week
        start = date.fromisoformat(report["start_date"])
        end = date.fromisoformat(report["end_date"])
        assert (end - start).days == 6

    def test_generate_monthly_report(self, report_generator, sample_employee):
        """Test monthly report generation"""
        report = report_generator.generate_report(
            sample_employee.id,
            ReportType.MONTHLY
        )

        assert report["report_type"] == ReportType.MONTHLY.value
        # Verify start is first day of month
        start = date.fromisoformat(report["start_date"])
        assert start.day == 1

    def test_generate_quarterly_report(
        self,
        report_generator,
        sample_employee
    ):
        """Test quarterly report generation"""
        report = report_generator.generate_report(
            sample_employee.id,
            ReportType.QUARTERLY
        )

        assert report["report_type"] == ReportType.QUARTERLY.value
        # Verify date range is approximately 3 months
        start = date.fromisoformat(report["start_date"])
        end = date.fromisoformat(report["end_date"])
        assert (end - start).days >= 89  # Minimum quarter length

    def test_generate_yearly_report(self, report_generator, sample_employee):
        """Test yearly report generation"""
        report = report_generator.generate_report(
            sample_employee.id,
            ReportType.YEARLY
        )

        assert report["report_type"] == ReportType.YEARLY.value
        # Verify start is Jan 1 and end is Dec 31
        start = date.fromisoformat(report["start_date"])
        end = date.fromisoformat(report["end_date"])
        assert start.month == 1 and start.day == 1
        assert end.month == 12 and end.day == 31

    def test_generate_since_start_report(
        self,
        report_generator,
        sample_employee
    ):
        """Test since start report generation"""
        report = report_generator.generate_report(
            sample_employee.id,
            ReportType.SINCE_START
        )

        assert report["report_type"] == ReportType.SINCE_START.value
        # Verify start date is employee's start date
        start = date.fromisoformat(report["start_date"])
        assert start == sample_employee.start_date

    def test_generate_report_with_custom_dates(
        self,
        report_generator,
        sample_employee
    ):
        """Test report generation with custom date range"""
        start_date = date(2023, 1, 1)
        end_date = date(2023, 12, 31)

        report = report_generator.generate_report(
            sample_employee.id,
            ReportType.DAILY,
            start_date=start_date,
            end_date=end_date
        )

        assert report["start_date"] == start_date.isoformat()
        assert report["end_date"] == end_date.isoformat()

    def test_generate_report_invalid_employee(self, report_generator):
        """Test report generation with invalid employee ID"""
        with pytest.raises(ValueError, match="not found"):
            report_generator.generate_report(
                99999,
                ReportType.DAILY
            )


class TestComparisonReports:
    """Test comparison report generation"""

    def test_generate_comparison_report(
        self,
        report_generator,
        db_session,
        sample_employee
    ):
        """Test comparison report with multiple employees"""
        # Create second employee
        emp_mgr = EmployeeManager(db_session)
        employee2 = emp_mgr.create_employee(
            first_name="Jane",
            last_name="Smith",
            city="Munich",
            birth_date=date(1992, 5, 15),
            position_id=sample_employee.position_id,
            start_date=date(2023, 6, 1)
        )

        # Generate comparison report
        report = report_generator.generate_comparison_report(
            [sample_employee.id, employee2.id],
            ReportType.MONTHLY
        )

        assert report["employee_count"] == 2
        assert len(report["employee_reports"]) == 2
        assert report["employee_ids"] == [sample_employee.id, employee2.id]

    def test_comparison_report_too_many_employees(
        self,
        report_generator,
        sample_employee
    ):
        """Test comparison report with more than 10 employees"""
        employee_ids = list(range(1, 12))  # 11 employees

        with pytest.raises(ValueError, match="maximum 10 employees"):
            report_generator.generate_comparison_report(
                employee_ids,
                ReportType.MONTHLY
            )

    def test_comparison_report_too_few_employees(
        self,
        report_generator,
        sample_employee
    ):
        """Test comparison report with less than 2 employees"""
        with pytest.raises(ValueError, match="at least 2 employees"):
            report_generator.generate_comparison_report(
                [sample_employee.id],
                ReportType.MONTHLY
            )


class TestReportSaving:
    """Test report saving and loading"""

    def test_save_and_load_report(
        self,
        report_generator,
        sample_employee
    ):
        """Test saving and loading a report"""
        # Generate report
        report = report_generator.generate_report(
            sample_employee.id,
            ReportType.DAILY
        )

        # Save report
        report_id = report_generator.save_report(report)
        assert report_id > 0

        # Load report
        loaded_report = report_generator.load_report(report_id)

        # Verify data matches
        assert loaded_report["employee_id"] == report["employee_id"]
        assert loaded_report["report_type"] == report["report_type"]
        assert loaded_report["start_date"] == report["start_date"]
        assert loaded_report["end_date"] == report["end_date"]
        assert "report_id" in loaded_report
        assert "created_at" in loaded_report

    def test_save_comparison_report(
        self,
        report_generator,
        db_session,
        sample_employee
    ):
        """Test saving a comparison report"""
        # Create second employee
        emp_mgr = EmployeeManager(db_session)
        employee2 = emp_mgr.create_employee(
            first_name="Bob",
            last_name="Johnson",
            city="Hamburg",
            birth_date=date(1988, 3, 20),
            position_id=sample_employee.position_id,
            start_date=date(2023, 3, 1)
        )

        # Generate comparison report
        report = report_generator.generate_comparison_report(
            [sample_employee.id, employee2.id],
            ReportType.WEEKLY
        )

        # Save report
        report_id = report_generator.save_report(report, is_comparison=True)
        assert report_id > 0

        # Load report
        loaded_report = report_generator.load_report(report_id)
        assert loaded_report["employee_count"] == 2

    def test_load_nonexistent_report(self, report_generator):
        """Test loading a report that doesn't exist"""
        with pytest.raises(ValueError, match="not found"):
            report_generator.load_report(99999)


class TestReportListing:
    """Test report listing and filtering"""

    def test_list_all_reports(self, report_generator, sample_employee):
        """Test listing all reports"""
        # Generate and save multiple reports
        for report_type in [ReportType.DAILY, ReportType.WEEKLY]:
            report = report_generator.generate_report(
                sample_employee.id,
                report_type
            )
            report_generator.save_report(report)

        # List reports
        reports = report_generator.list_reports()
        assert len(reports) >= 2

        # Verify reports are sorted by creation date (newest first)
        for i in range(len(reports) - 1):
            assert reports[i]["created_at"] >= reports[i + 1]["created_at"]

    def test_list_reports_by_employee(
        self,
        report_generator,
        db_session,
        sample_employee
    ):
        """Test filtering reports by employee"""
        # Create second employee
        emp_mgr = EmployeeManager(db_session)
        employee2 = emp_mgr.create_employee(
            first_name="Alice",
            last_name="Williams",
            city="Frankfurt",
            birth_date=date(1995, 7, 10),
            position_id=sample_employee.position_id,
            start_date=date(2023, 9, 1)
        )

        # Generate reports for both employees
        report1 = report_generator.generate_report(
            sample_employee.id,
            ReportType.DAILY
        )
        report_generator.save_report(report1)

        report2 = report_generator.generate_report(
            employee2.id,
            ReportType.DAILY
        )
        report_generator.save_report(report2)

        # List reports for first employee only
        reports = report_generator.list_reports(
            employee_id=sample_employee.id
        )

        # Verify only first employee's reports are returned
        for report in reports:
            if report["employee_id"] is not None:
                assert report["employee_id"] == sample_employee.id

    def test_list_reports_by_type(self, report_generator, sample_employee):
        """Test filtering reports by type"""
        # Generate reports of different types
        for report_type in [ReportType.DAILY, ReportType.WEEKLY]:
            report = report_generator.generate_report(
                sample_employee.id,
                report_type
            )
            report_generator.save_report(report)

        # List only daily reports
        reports = report_generator.list_reports(report_type=ReportType.DAILY)

        # Verify only daily reports are returned
        for report in reports:
            assert report["report_type"] == ReportType.DAILY.value

    def test_list_reports_with_limit(
        self,
        report_generator,
        sample_employee
    ):
        """Test limiting number of reports returned"""
        # Generate multiple reports
        for _ in range(5):
            report = report_generator.generate_report(
                sample_employee.id,
                ReportType.DAILY
            )
            report_generator.save_report(report)

        # List with limit
        reports = report_generator.list_reports(limit=3)
        assert len(reports) <= 3


class TestReportExport:
    """Test report export functionality"""

    def test_export_json(self, report_generator, sample_employee):
        """Test JSON export"""
        # Generate report
        report = report_generator.generate_report(
            sample_employee.id,
            ReportType.DAILY
        )

        # Export to JSON
        json_str = report_generator.export_report_json(report)

        # Verify it's valid JSON
        parsed = json.loads(json_str)
        assert "report_metadata" in parsed
        assert "quotas" in parsed
        assert "ratio_descriptions" in parsed
        assert parsed["report_metadata"]["employee_id"] == sample_employee.id

    def test_export_json_round_trip(self, report_generator, sample_employee):
        """Test JSON export and parse round-trip"""
        # Generate report
        report = report_generator.generate_report(
            sample_employee.id,
            ReportType.WEEKLY
        )

        # Export to JSON
        json_str = report_generator.export_report_json(report)

        # Parse back
        parsed = json.loads(json_str)

        # Verify key data is preserved
        assert (
            parsed["report_metadata"]["employee_name"] ==
            report["employee_name"]
        )
        assert (
            parsed["report_metadata"]["report_type"] ==
            report["report_type"]
        )
        assert len(parsed["quotas"]) == len(report["quotas"])

    def test_export_excel(self, report_generator, sample_employee):
        """Test Excel export"""
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            pytest.skip("openpyxl not installed")

        # Generate report
        report = report_generator.generate_report(
            sample_employee.id,
            ReportType.MONTHLY
        )

        # Export to Excel
        excel_bytes = report_generator.export_report_excel(report)

        # Verify it's valid Excel data
        assert isinstance(excel_bytes, bytes)
        assert len(excel_bytes) > 0

        # Try to load it with openpyxl
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb.active

        # Verify some content
        assert ws["A1"].value == "Employee Controlling Report"

    def test_export_pdf(self, report_generator, sample_employee):
        """Test PDF export"""
        try:
            import reportlab  # noqa: F401
        except ImportError:
            pytest.skip("reportlab not installed")

        # Generate report
        report = report_generator.generate_report(
            sample_employee.id,
            ReportType.QUARTERLY
        )

        # Export to PDF
        pdf_bytes = report_generator.export_report_pdf(report)

        # Verify it's valid PDF data
        assert isinstance(pdf_bytes, bytes)
        assert len(pdf_bytes) > 0
        # PDF files start with %PDF
        assert pdf_bytes[:4] == b'%PDF'

    def test_export_all_formats(self, report_generator, sample_employee):
        """Test that all export formats are available"""
        # Generate report
        report = report_generator.generate_report(
            sample_employee.id,
            ReportType.YEARLY
        )

        # JSON should always work
        json_str = report_generator.export_report_json(report)
        assert json_str is not None

        # Excel and PDF depend on libraries
        try:
            excel_bytes = report_generator.export_report_excel(report)
            assert excel_bytes is not None
        except ImportError:
            pass  # OK if library not installed

        try:
            pdf_bytes = report_generator.export_report_pdf(report)
            assert pdf_bytes is not None
        except ImportError:
            pass  # OK if library not installed

    def test_export_comparison_report_json(
        self,
        report_generator,
        db_session,
        sample_employee
    ):
        """Test JSON export of comparison report"""
        # Create second employee
        emp_mgr = EmployeeManager(db_session)
        employee2 = emp_mgr.create_employee(
            first_name="Export",
            last_name="Test",
            city="Berlin",
            birth_date=date(1991, 6, 15),
            position_id=sample_employee.position_id,
            start_date=date(2023, 4, 1)
        )

        # Generate comparison report
        report = report_generator.generate_comparison_report(
            [sample_employee.id, employee2.id],
            ReportType.MONTHLY
        )

        # Export to JSON
        json_str = report_generator.export_report_json(report)
        parsed = json.loads(json_str)

        # Verify comparison data
        assert "employee_reports" in parsed
        assert len(parsed["employee_reports"]) == 2
