"""
Excel Integration - Import Module

Dieses Modul implementiert Import-Funktionalität für CSV und Excel-Dateien.
"""

import csv
import io
import chardet
from typing import Optional, Tuple, List, Dict, Any

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from excel.excel_models import ExcelMatrix, Cell
from excel.excel_manager import ExcelManager


class ImportError(Exception):
    """Fehler beim Import"""
    pass


def detect_encoding(file_content: bytes) -> str:
    """
    Erkennt das Encoding einer Datei
    
    Args:
        file_content: Dateiinhalt als Bytes
        
    Returns:
        Encoding-String (z.B. 'utf-8', 'latin-1')
    """
    result = chardet.detect(file_content)
    encoding = result.get('encoding', 'utf-8')
    
    # Fallback auf utf-8 wenn Erkennung unsicher
    if not encoding or result.get('confidence', 0) < 0.7:
        encoding = 'utf-8'
    
    return encoding


def detect_delimiter(csv_text: str, sample_lines: int = 5) -> str:
    """
    Erkennt den Delimiter einer CSV-Datei
    
    Args:
        csv_text: CSV-Text
        sample_lines: Anzahl der Zeilen für die Analyse
        
    Returns:
        Delimiter-Zeichen (';', ',', oder '\t')
    """
    # Nimm erste paar Zeilen als Sample
    lines = csv_text.split('\n')[:sample_lines]
    sample = '\n'.join(lines)
    
    # Zähle Vorkommen verschiedener Delimiter
    delimiters = [';', ',', '\t', '|']
    counts = {}
    
    for delimiter in delimiters:
        # Zähle wie oft der Delimiter in jeder Zeile vorkommt
        line_counts = []
        for line in lines:
            if line.strip():
                line_counts.append(line.count(delimiter))
        
        # Prüfe ob Delimiter konsistent verwendet wird
        if line_counts and len(set(line_counts)) == 1 and line_counts[0] > 0:
            counts[delimiter] = line_counts[0]
    
    # Wähle Delimiter mit den meisten Vorkommen
    if counts:
        return max(counts.items(), key=lambda x: x[1])[0]
    
    # Fallback auf Semikolon (Standard in Deutschland)
    return ';'


def parse_csv_content(
    csv_text: str,
    delimiter: Optional[str] = None,
    has_header: bool = True
) -> Tuple[List[str], List[List[str]]]:
    """
    Parst CSV-Inhalt
    
    Args:
        csv_text: CSV-Text
        delimiter: Optionaler Delimiter (wird automatisch erkannt wenn None)
        has_header: Ob erste Zeile Header ist
        
    Returns:
        Tupel (header, rows) mit Header-Liste und Daten-Zeilen
        
    Raises:
        ImportError: Bei Parsing-Fehlern
    """
    try:
        # Delimiter erkennen wenn nicht angegeben
        if delimiter is None:
            delimiter = detect_delimiter(csv_text)
        
        # Parse CSV
        reader = csv.reader(io.StringIO(csv_text), delimiter=delimiter)
        rows = list(reader)
        
        if not rows:
            raise ImportError("CSV-Datei ist leer")
        
        # Extrahiere Header
        if has_header:
            header = rows[0]
            data_rows = rows[1:]
        else:
            # Generiere automatische Header (A, B, C, ...)
            num_cols = len(rows[0]) if rows else 0
            header = [chr(65 + i) for i in range(num_cols)]
            data_rows = rows
        
        return header, data_rows
        
    except Exception as e:
        raise ImportError(f"Fehler beim Parsen der CSV-Datei: {str(e)}")


def import_csv_to_matrix(
    file_content: bytes,
    matrix_name: str,
    delimiter: Optional[str] = None,
    has_header: bool = True,
    encoding: Optional[str] = None
) -> ExcelManager:
    """
    Importiert eine CSV-Datei als neue Excel-Matrix
    
    Args:
        file_content: Dateiinhalt als Bytes
        matrix_name: Name für die neue Matrix
        delimiter: Optionaler Delimiter (wird automatisch erkannt wenn None)
        has_header: Ob erste Zeile Header ist
        encoding: Optionales Encoding (wird automatisch erkannt wenn None)
        
    Returns:
        ExcelManager mit importierter Matrix
        
    Raises:
        ImportError: Bei Fehlern während des Imports
    """
    try:
        # Encoding erkennen wenn nicht angegeben
        if encoding is None:
            encoding = detect_encoding(file_content)
        
        # Dekodiere Dateiinhalt
        try:
            csv_text = file_content.decode(encoding)
        except UnicodeDecodeError:
            # Fallback auf utf-8 mit Fehlerbehandlung
            csv_text = file_content.decode('utf-8', errors='replace')
        
        # Parse CSV
        header, data_rows = parse_csv_content(csv_text, delimiter, has_header)
        
        # Erstelle neue Matrix
        num_rows = len(data_rows)
        num_cols = len(header)
        
        matrix = ExcelMatrix(
            name=matrix_name,
            description=f"Importiert aus CSV",
            rows=max(num_rows, 100),  # Mindestens 100 Zeilen
            columns=max(num_cols, 26)  # Mindestens 26 Spalten (A-Z)
        )
        
        # Fülle Matrix mit Daten
        for row_idx, row_data in enumerate(data_rows):
            for col_idx, cell_value in enumerate(row_data):
                if col_idx >= num_cols:
                    break  # Ignoriere überzählige Spalten
                
                # Überspringe leere Zellen
                if not cell_value or cell_value.strip() == '':
                    continue
                
                # Erstelle Zelle
                cell = Cell(row=row_idx, col=col_idx)
                
                # Prüfe ob Formel
                if cell_value.startswith('='):
                    cell.formula = cell_value
                    cell.raw_input = cell_value
                    cell.data_type = "formula"
                else:
                    # Versuche als Zahl zu parsen
                    try:
                        # Ersetze Komma durch Punkt für deutsche Zahlen
                        numeric_value = float(cell_value.replace(',', '.'))
                        cell.value = numeric_value
                        cell.data_type = "number"
                    except ValueError:
                        # Behandle als Text
                        cell.value = cell_value
                        cell.data_type = "text"
                    
                    cell.raw_input = cell_value
                
                # Füge Zelle zur Matrix hinzu
                matrix.cells[(row_idx, col_idx)] = cell
        
        # Erstelle ExcelManager
        manager = ExcelManager(matrix)
        
        # Berechne alle Formeln
        manager.recalculate_all_formulas()
        
        return manager
        
    except ImportError:
        raise
    except Exception as e:
        raise ImportError(f"Fehler beim Import: {str(e)}")


def validate_csv_file(file_content: bytes) -> Dict[str, Any]:
    """
    Validiert eine CSV-Datei und gibt Informationen zurück
    
    Args:
        file_content: Dateiinhalt als Bytes
        
    Returns:
        Dictionary mit Validierungsinformationen:
        {
            'valid': bool,
            'encoding': str,
            'delimiter': str,
            'num_rows': int,
            'num_cols': int,
            'has_formulas': bool,
            'errors': List[str]
        }
    """
    result = {
        'valid': False,
        'encoding': None,
        'delimiter': None,
        'num_rows': 0,
        'num_cols': 0,
        'has_formulas': False,
        'errors': []
    }
    
    try:
        # Encoding erkennen
        encoding = detect_encoding(file_content)
        result['encoding'] = encoding
        
        # Dekodiere
        try:
            csv_text = file_content.decode(encoding)
        except UnicodeDecodeError as e:
            result['errors'].append(f"Encoding-Fehler: {str(e)}")
            return result
        
        # Delimiter erkennen
        delimiter = detect_delimiter(csv_text)
        result['delimiter'] = delimiter
        
        # Parse CSV
        try:
            header, data_rows = parse_csv_content(csv_text, delimiter)
            result['num_rows'] = len(data_rows)
            result['num_cols'] = len(header)
            
            # Prüfe auf Formeln
            for row in data_rows:
                for cell in row:
                    if cell.startswith('='):
                        result['has_formulas'] = True
                        break
                if result['has_formulas']:
                    break
            
            result['valid'] = True
            
        except Exception as e:
            result['errors'].append(f"Parse-Fehler: {str(e)}")
            return result
        
    except Exception as e:
        result['errors'].append(f"Validierungsfehler: {str(e)}")
    
    return result


def get_csv_preview(
    file_content: bytes,
    max_rows: int = 10,
    encoding: Optional[str] = None,
    delimiter: Optional[str] = None
) -> Dict[str, Any]:
    """
    Erstellt eine Vorschau einer CSV-Datei
    
    Args:
        file_content: Dateiinhalt als Bytes
        max_rows: Maximale Anzahl Zeilen für Vorschau
        encoding: Optionales Encoding
        delimiter: Optionaler Delimiter
        
    Returns:
        Dictionary mit Vorschau-Daten:
        {
            'header': List[str],
            'rows': List[List[str]],
            'total_rows': int,
            'encoding': str,
            'delimiter': str
        }
    """
    # Encoding erkennen
    if encoding is None:
        encoding = detect_encoding(file_content)
    
    # Dekodiere
    csv_text = file_content.decode(encoding, errors='replace')
    
    # Delimiter erkennen
    if delimiter is None:
        delimiter = detect_delimiter(csv_text)
    
    # Parse CSV
    header, data_rows = parse_csv_content(csv_text, delimiter)
    
    # Begrenze auf max_rows
    preview_rows = data_rows[:max_rows]
    
    return {
        'header': header,
        'rows': preview_rows,
        'total_rows': len(data_rows),
        'encoding': encoding,
        'delimiter': delimiter
    }



# ============================================================================
# Excel (XLS/XLSX) Import Functions
# ============================================================================


def get_excel_sheet_names(file_content: bytes) -> List[str]:
    """
    Gibt die Namen aller Sheets in einer Excel-Datei zurück

    Args:
        file_content: Dateiinhalt als Bytes

    Returns:
        Liste der Sheet-Namen

    Raises:
        ImportError: Wenn openpyxl nicht verfügbar ist
        Exception: Bei Fehlern beim Lesen der Datei
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "openpyxl ist nicht installiert. "
            "Bitte installieren Sie es mit: pip install openpyxl"
        )

    try:
        # Lade Workbook aus Bytes
        wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        return sheet_names
    except Exception as e:
        raise ImportError(f"Fehler beim Lesen der Excel-Datei: {str(e)}")


def parse_excel_content(
    file_content: bytes,
    sheet_name: Optional[str] = None,
    has_header: bool = True
) -> Tuple[List[str], List[List[Any]]]:
    """
    Parst Excel-Inhalt und extrahiert Daten

    Args:
        file_content: Dateiinhalt als Bytes
        sheet_name: Optionaler Sheet-Name (None = erstes Sheet)
        has_header: Ob erste Zeile Header ist

    Returns:
        Tupel (header, rows) mit Header-Liste und Daten-Zeilen

    Raises:
        ImportError: Bei Parsing-Fehlern oder wenn openpyxl fehlt
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "openpyxl ist nicht installiert. "
            "Bitte installieren Sie es mit: pip install openpyxl"
        )

    try:
        # Lade Workbook (data_only=False um Formeln zu erhalten)
        wb = openpyxl.load_workbook(
            io.BytesIO(file_content),
            data_only=False
        )

        # Wähle Sheet
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise ImportError(
                    f"Sheet '{sheet_name}' nicht gefunden. "
                    f"Verfügbare Sheets: {', '.join(wb.sheetnames)}"
                )
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # Extrahiere alle Zeilen
        all_rows = []
        for row in ws.iter_rows(values_only=False):
            row_data = []
            for cell in row:
                # Prüfe ob Formel
                if cell.data_type == 'f':
                    # Speichere Formel mit '=' (falls nicht schon vorhanden)
                    formula = cell.value
                    if formula and not formula.startswith('='):
                        formula = f"={formula}"
                    row_data.append(formula)
                else:
                    # Normaler Wert
                    row_data.append(cell.value)
            all_rows.append(row_data)

        wb.close()

        if not all_rows:
            raise ImportError("Excel-Datei ist leer")

        # Extrahiere Header
        if has_header:
            header = [str(h) if h is not None else "" for h in all_rows[0]]
            data_rows = all_rows[1:]
        else:
            # Generiere automatische Header (A, B, C, ...)
            num_cols = len(all_rows[0]) if all_rows else 0
            header = [get_column_letter(i + 1) for i in range(num_cols)]
            data_rows = all_rows

        return header, data_rows

    except ImportError:
        raise
    except Exception as e:
        raise ImportError(f"Fehler beim Parsen der Excel-Datei: {str(e)}")


def import_excel_to_matrix(
    file_content: bytes,
    matrix_name: str,
    sheet_name: Optional[str] = None,
    has_header: bool = True
) -> ExcelManager:
    """
    Importiert eine Excel-Datei (XLS/XLSX) als neue Excel-Matrix

    Args:
        file_content: Dateiinhalt als Bytes
        matrix_name: Name für die neue Matrix
        sheet_name: Optionaler Sheet-Name (None = erstes Sheet)
        has_header: Ob erste Zeile Header ist

    Returns:
        ExcelManager mit importierter Matrix

    Raises:
        ImportError: Bei Fehlern während des Imports
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "openpyxl ist nicht installiert. "
            "Bitte installieren Sie es mit: pip install openpyxl"
        )

    try:
        # Parse Excel
        header, data_rows = parse_excel_content(
            file_content,
            sheet_name,
            has_header
        )

        # Erstelle neue Matrix
        num_rows = len(data_rows)
        num_cols = len(header)

        matrix = ExcelMatrix(
            name=matrix_name,
            description=f"Importiert aus Excel{f' (Sheet: {sheet_name})' if sheet_name else ''}",
            rows=max(num_rows, 100),  # Mindestens 100 Zeilen
            columns=max(num_cols, 26)  # Mindestens 26 Spalten (A-Z)
        )

        # Fülle Matrix mit Daten
        for row_idx, row_data in enumerate(data_rows):
            for col_idx, cell_value in enumerate(row_data):
                if col_idx >= num_cols:
                    break  # Ignoriere überzählige Spalten

                # Überspringe leere Zellen
                if cell_value is None or (isinstance(cell_value, str) and cell_value.strip() == ''):
                    continue

                # Erstelle Zelle
                cell = Cell(row=row_idx, col=col_idx)

                # Prüfe ob Formel
                if isinstance(cell_value, str) and cell_value.startswith('='):
                    cell.formula = cell_value
                    cell.raw_input = cell_value
                    cell.data_type = "formula"
                else:
                    # Versuche als Zahl zu parsen
                    if isinstance(cell_value, (int, float)):
                        cell.value = float(cell_value)
                        cell.data_type = "number"
                    elif isinstance(cell_value, str):
                        try:
                            # Ersetze Komma durch Punkt für deutsche Zahlen
                            numeric_value = float(cell_value.replace(',', '.'))
                            cell.value = numeric_value
                            cell.data_type = "number"
                        except ValueError:
                            # Behandle als Text
                            cell.value = cell_value
                            cell.data_type = "text"
                    else:
                        # Andere Typen (Datum, etc.) als Text
                        cell.value = str(cell_value)
                        cell.data_type = "text"

                    cell.raw_input = str(cell_value)

                # Füge Zelle zur Matrix hinzu
                matrix.cells[(row_idx, col_idx)] = cell

        # Erstelle ExcelManager
        manager = ExcelManager(matrix)

        # Berechne alle Formeln
        manager.recalculate_all_formulas()

        return manager

    except ImportError:
        raise
    except Exception as e:
        raise ImportError(f"Fehler beim Import: {str(e)}")


def validate_excel_file(file_content: bytes) -> Dict[str, Any]:
    """
    Validiert eine Excel-Datei und gibt Informationen zurück

    Args:
        file_content: Dateiinhalt als Bytes

    Returns:
        Dictionary mit Validierungsinformationen:
        {
            'valid': bool,
            'sheets': List[str],
            'num_rows': int,
            'num_cols': int,
            'has_formulas': bool,
            'errors': List[str]
        }
    """
    result = {
        'valid': False,
        'sheets': [],
        'num_rows': 0,
        'num_cols': 0,
        'has_formulas': False,
        'errors': []
    }

    if not OPENPYXL_AVAILABLE:
        result['errors'].append(
            "openpyxl ist nicht installiert. "
            "Bitte installieren Sie es mit: pip install openpyxl"
        )
        return result

    try:
        # Hole Sheet-Namen
        try:
            sheet_names = get_excel_sheet_names(file_content)
            result['sheets'] = sheet_names
        except Exception as e:
            result['errors'].append(f"Fehler beim Lesen der Sheets: {str(e)}")
            return result

        # Parse erstes Sheet
        try:
            header, data_rows = parse_excel_content(file_content)
            result['num_rows'] = len(data_rows)
            result['num_cols'] = len(header)

            # Prüfe auf Formeln
            for row in data_rows:
                for cell in row:
                    if isinstance(cell, str) and cell.startswith('='):
                        result['has_formulas'] = True
                        break
                if result['has_formulas']:
                    break

            result['valid'] = True

        except Exception as e:
            result['errors'].append(f"Parse-Fehler: {str(e)}")
            return result

    except Exception as e:
        result['errors'].append(f"Validierungsfehler: {str(e)}")

    return result


def get_excel_preview(
    file_content: bytes,
    max_rows: int = 10,
    sheet_name: Optional[str] = None
) -> Dict[str, Any]:
    """
    Erstellt eine Vorschau einer Excel-Datei

    Args:
        file_content: Dateiinhalt als Bytes
        max_rows: Maximale Anzahl Zeilen für Vorschau
        sheet_name: Optionaler Sheet-Name

    Returns:
        Dictionary mit Vorschau-Daten:
        {
            'header': List[str],
            'rows': List[List[Any]],
            'total_rows': int,
            'sheets': List[str],
            'current_sheet': str
        }

    Raises:
        ImportError: Wenn openpyxl nicht verfügbar ist
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "openpyxl ist nicht installiert. "
            "Bitte installieren Sie es mit: pip install openpyxl"
        )

    # Hole Sheet-Namen
    sheet_names = get_excel_sheet_names(file_content)

    # Parse Excel
    header, data_rows = parse_excel_content(file_content, sheet_name)

    # Begrenze auf max_rows
    preview_rows = data_rows[:max_rows]

    # Konvertiere Werte zu Strings für Anzeige
    preview_rows_str = []
    for row in preview_rows:
        row_str = []
        for cell in row:
            if cell is None:
                row_str.append("")
            else:
                row_str.append(str(cell))
        preview_rows_str.append(row_str)

    return {
        'header': header,
        'rows': preview_rows_str,
        'total_rows': len(data_rows),
        'sheets': sheet_names,
        'current_sheet': sheet_name or sheet_names[0] if sheet_names else None
    }
