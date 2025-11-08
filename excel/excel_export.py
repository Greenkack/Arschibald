"""
Excel Integration - Export Module

Dieses Modul implementiert Export-Funktionalität für CSV und Excel-Dateien.
"""

import csv
import io
from typing import Optional, Dict, Any
from datetime import datetime

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from excel.excel_manager import ExcelManager


class ExportError(Exception):
    """Fehler beim Export"""
    pass


def export_to_csv(
    manager: ExcelManager,
    delimiter: str = ';',
    include_formulas: bool = False,
    encoding: str = 'utf-8'
) -> bytes:
    """
    Exportiert eine Matrix als CSV-Datei

    Args:
        manager: ExcelManager mit der zu exportierenden Matrix
        delimiter: CSV-Delimiter (Standard: ';')
        include_formulas: Ob Formeln statt Werte exportiert werden sollen
        encoding: Encoding für die CSV-Datei (Standard: 'utf-8')

    Returns:
        CSV-Datei als Bytes

    Raises:
        ExportError: Bei Fehlern während des Exports
    """
    try:
        matrix = manager.get_matrix()

        # Erstelle CSV in Memory
        output = io.StringIO()
        writer = csv.writer(output, delimiter=delimiter)

        # Schreibe Header (Spaltenbezeichnungen)
        header = ['']  # Erste Spalte für Zeilennummern
        for col_idx in range(matrix.columns):
            col_letter = get_column_letter(col_idx + 1)
            header.append(col_letter)
        writer.writerow(header)

        # Schreibe Daten
        for row_idx in range(matrix.rows):
            row_data = [str(row_idx + 1)]  # Zeilennummer

            for col_idx in range(matrix.columns):
                cell = matrix.get_cell(row_idx, col_idx)

                if cell.value is None and not cell.is_formula():
                    # Leere Zelle
                    row_data.append('')
                elif include_formulas and cell.is_formula():
                    # Exportiere Formel
                    row_data.append(cell.formula)
                else:
                    # Exportiere Wert
                    value = cell.get_display_value()
                    row_data.append(value)

            writer.writerow(row_data)

        # Konvertiere zu Bytes
        csv_text = output.getvalue()
        return csv_text.encode(encoding)

    except Exception as e:
        raise ExportError(f"Fehler beim CSV-Export: {str(e)}")


def export_to_excel(
    manager: ExcelManager,
    include_formulas: bool = True,
    include_formatting: bool = True
) -> bytes:
    """
    Exportiert eine Matrix als Excel-Datei (XLSX) mit Formeln

    Args:
        manager: ExcelManager mit der zu exportierenden Matrix
        include_formulas: Ob Formeln exportiert werden sollen
        include_formatting: Ob Formatierung angewendet werden soll

    Returns:
        Excel-Datei als Bytes

    Raises:
        ExportError: Bei Fehlern während des Exports oder wenn openpyxl fehlt
    """
    if not OPENPYXL_AVAILABLE:
        raise ExportError(
            "openpyxl ist nicht installiert. "
            "Bitte installieren Sie es mit: pip install openpyxl"
        )

    try:
        matrix = manager.get_matrix()

        # Erstelle Workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = matrix.name[:31]  # Excel limit für Sheet-Namen

        # Formatierung für Header
        if include_formatting:
            header_fill = PatternFill(
                start_color="366092",
                end_color="366092",
                fill_type="solid"
            )
            header_font = Font(color="FFFFFF", bold=True)
            header_alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

        # Schreibe Header (Spaltenbezeichnungen)
        ws.cell(1, 1, "")  # Erste Zelle leer
        for col_idx in range(matrix.columns):
            col_letter = get_column_letter(col_idx + 1)
            cell = ws.cell(1, col_idx + 2, col_letter)

            if include_formatting:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

        # Schreibe Zeilennummern und Daten
        for row_idx in range(matrix.rows):
            # Zeilennummer in erste Spalte
            row_cell = ws.cell(row_idx + 2, 1, row_idx + 1)

            if include_formatting:
                row_cell.fill = header_fill
                row_cell.font = header_font
                row_cell.alignment = header_alignment

            # Zellwerte
            for col_idx in range(matrix.columns):
                matrix_cell = matrix.get_cell(row_idx, col_idx)
                excel_cell = ws.cell(row_idx + 2, col_idx + 2)

                if matrix_cell.value is None and not matrix_cell.is_formula():
                    # Leere Zelle
                    excel_cell.value = None
                elif include_formulas and matrix_cell.is_formula():
                    # Exportiere Formel
                    # Entferne führendes '=' wenn vorhanden
                    formula = matrix_cell.formula
                    if formula.startswith('='):
                        formula = formula[1:]
                    excel_cell.value = f"={formula}"
                else:
                    # Exportiere Wert
                    if matrix_cell.data_type == "number":
                        try:
                            excel_cell.value = float(matrix_cell.value)
                        except (ValueError, TypeError):
                            excel_cell.value = matrix_cell.get_display_value()
                    else:
                        excel_cell.value = matrix_cell.get_display_value()

                # Formatierung für Fehler
                if include_formatting and matrix_cell.is_error():
                    excel_cell.fill = PatternFill(
                        start_color="FFC7CE",
                        end_color="FFC7CE",
                        fill_type="solid"
                    )
                    excel_cell.font = Font(color="9C0006")

        # Passe Spaltenbreiten an
        if include_formatting:
            for col_idx in range(matrix.columns + 1):
                col_letter = get_column_letter(col_idx + 1)
                ws.column_dimensions[col_letter].width = 12

        # Speichere in Memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()

    except ExportError:
        raise
    except Exception as e:
        raise ExportError(f"Fehler beim Excel-Export: {str(e)}")


def generate_filename(
    matrix_name: str,
    file_format: str,
    include_timestamp: bool = True
) -> str:
    """
    Generiert einen Dateinamen für den Export

    Args:
        matrix_name: Name der Matrix
        file_format: Dateiformat ('csv' oder 'xlsx')
        include_timestamp: Ob Zeitstempel hinzugefügt werden soll

    Returns:
        Generierter Dateiname

    Raises:
        ValueError: Bei ungültigem Dateiformat
    """
    if file_format not in ['csv', 'xlsx']:
        raise ValueError(
            f"Ungültiges Dateiformat: {file_format}. "
            "Erlaubt sind: 'csv', 'xlsx'"
        )

    # Bereinige Matrix-Namen (entferne ungültige Zeichen)
    safe_name = "".join(
        c for c in matrix_name
        if c.isalnum() or c in (' ', '-', '_')
    ).strip()

    # Ersetze Leerzeichen durch Unterstriche
    safe_name = safe_name.replace(' ', '_')

    # Füge Zeitstempel hinzu wenn gewünscht
    if include_timestamp:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{safe_name}_{timestamp}.{file_format}"
    else:
        filename = f"{safe_name}.{file_format}"

    return filename


def get_export_info(manager: ExcelManager) -> Dict[str, Any]:
    """
    Gibt Informationen über den Export zurück

    Args:
        manager: ExcelManager mit der Matrix

    Returns:
        Dictionary mit Export-Informationen:
        {
            'matrix_name': str,
            'rows': int,
            'columns': int,
            'cell_count': int,
            'formula_count': int,
            'has_errors': bool,
            'estimated_csv_size': int,  # in Bytes
            'estimated_xlsx_size': int  # in Bytes
        }
    """
    matrix = manager.get_matrix()
    info = manager.get_matrix_info()

    # Zähle Fehler
    error_count = sum(
        1 for cell in matrix.cells.values()
        if cell.is_error()
    )

    # Schätze Dateigrößen (grobe Schätzung)
    # CSV: ~10 Bytes pro Zelle
    estimated_csv_size = info['cell_count'] * 10

    # XLSX: ~50 Bytes pro Zelle + Overhead
    estimated_xlsx_size = info['cell_count'] * 50 + 5000

    return {
        'matrix_name': matrix.name,
        'rows': info['rows'],
        'columns': info['columns'],
        'cell_count': info['cell_count'],
        'formula_count': info['formula_count'],
        'has_errors': error_count > 0,
        'error_count': error_count,
        'estimated_csv_size': estimated_csv_size,
        'estimated_xlsx_size': estimated_xlsx_size
    }


def validate_export(manager: ExcelManager) -> Dict[str, Any]:
    """
    Validiert ob ein Export möglich ist

    Args:
        manager: ExcelManager mit der Matrix

    Returns:
        Dictionary mit Validierungsergebnis:
        {
            'valid': bool,
            'warnings': List[str],
            'errors': List[str]
        }
    """
    result = {
        'valid': True,
        'warnings': [],
        'errors': []
    }

    matrix = manager.get_matrix()
    info = manager.get_matrix_info()

    # Prüfe ob Matrix leer ist
    if info['cell_count'] == 0:
        result['warnings'].append(
            "Matrix ist leer. Export enthält nur Header."
        )

    # Prüfe auf Fehler in Formeln
    error_cells = [
        cell for cell in matrix.cells.values()
        if cell.is_error()
    ]

    if error_cells:
        result['warnings'].append(
            f"{len(error_cells)} Zelle(n) enthalten Fehler. "
            "Diese werden als Fehlerwerte exportiert."
        )

    # Prüfe Größe
    if info['cell_count'] > 100000:
        result['warnings'].append(
            "Matrix ist sehr groß (>100.000 Zellen). "
            "Export kann einige Zeit dauern."
        )

    # Prüfe auf sehr lange Werte
    long_values = [
        cell for cell in matrix.cells.values()
        if cell.value and len(str(cell.value)) > 1000
    ]

    if long_values:
        result['warnings'].append(
            f"{len(long_values)} Zelle(n) enthalten sehr lange Werte. "
            "Dies kann die Dateigröße erhöhen."
        )

    return result
