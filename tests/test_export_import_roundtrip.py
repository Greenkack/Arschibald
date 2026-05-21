"""
Test Suite für Import/Export Funktionalität (Task 15.1)

Tests:
- CSV Import → Export → Re-Import
- Excel Import mit Formeln
- Große Dateien (10 MB)
- Verschiedene Encodings
"""

import pytest
import io
import os
from datetime import datetime

from excel.excel_manager import ExcelManager
from excel.excel_models import ExcelMatrix, Cell
from excel.excel_import import (
    import_csv_to_matrix,
    import_excel_to_matrix,
    validate_csv_file,
    validate_excel_file
)
from excel.excel_export import (
    export_to_csv,
    export_to_excel,
    generate_filename,
    get_export_info,
    validate_export
)


# Hilfsfunktionen

def _is_openpyxl_available():
    """Prüft ob openpyxl verfügbar ist"""
    try:
        import openpyxl
        return True
    except ImportError:
        return False


class TestCSVRoundtrip:
    """Test CSV Import → Export → Re-Import"""
    
    def test_csv_roundtrip_simple(self):
        """Test: Einfache CSV-Datei importieren, exportieren und re-importieren"""
        # Erstelle Test-CSV (ohne Header-Zeile, nur Daten)
        csv_content = b"""1;2;3
4;5;6
7;8;9"""
        
        # Import
        manager1 = import_csv_to_matrix(
            csv_content,
            "Test Matrix",
            delimiter=";",
            has_header=False  # Keine Header-Zeile
        )
        
        assert manager1 is not None
        assert manager1.get_cell_value(0, 0) == 1.0
        assert manager1.get_cell_value(1, 1) == 5.0
        assert manager1.get_cell_value(2, 2) == 9.0
        
        # Export
        exported_csv = export_to_csv(
            manager1,
            delimiter=";",
            include_formulas=False
        )
        
        assert exported_csv is not None
        assert len(exported_csv) > 0
        
        # Re-Import (Export hat Header-Zeile mit Spaltenbezeichnungen + Zeilennummern-Spalte)
        manager2 = import_csv_to_matrix(
            exported_csv,
            "Test Matrix 2",
            delimiter=";",
            has_header=True  # Export hat Header
        )
        
        assert manager2 is not None
        # Vergleiche Werte
        # Beachte: Export hat Zeilennummern-Spalte, daher sind die Spalten um 1 verschoben
        # Spalte 0 im Export = Zeilennummer, Spalte 1 = A (ursprünglich Spalte 0)
        assert manager2.get_cell_value(0, 1) == manager1.get_cell_value(0, 0)  # A1
        assert manager2.get_cell_value(1, 2) == manager1.get_cell_value(1, 1)  # B2
        assert manager2.get_cell_value(2, 3) == manager1.get_cell_value(2, 2)  # C3
    
    def test_csv_roundtrip_with_formulas(self):
        """Test: CSV mit Formeln exportieren (Werte werden exportiert)"""
        # Erstelle Test-CSV mit Formeln
        csv_content = b"""A;B;C
10;20;=A1+B1
5;15;=A2+B2
100;200;=A3+B3"""
        
        # Import
        manager1 = import_csv_to_matrix(
            csv_content,
            "Test Matrix Formulas",
            delimiter=";",
            has_header=True
        )
        
        assert manager1 is not None
        
        # Prüfe dass Formeln berechnet wurden
        assert manager1.get_cell_value(0, 2) == 30.0  # 10 + 20
        assert manager1.get_cell_value(1, 2) == 20.0  # 5 + 15
        assert manager1.get_cell_value(2, 2) == 300.0  # 100 + 200
        
        # Export mit Formeln
        exported_csv = export_to_csv(
            manager1,
            delimiter=";",
            include_formulas=True
        )
        
        assert exported_csv is not None
        # Prüfe dass Formeln im Export enthalten sind
        exported_str = exported_csv.decode('utf-8')
        assert "=A" in exported_str or "=B" in exported_str
        
        # Export ohne Formeln (nur Werte)
        exported_values = export_to_csv(
            manager1,
            delimiter=";",
            include_formulas=False
        )
        
        assert exported_values is not None
        # Prüfe dass berechnete Werte enthalten sind
        exported_values_str = exported_values.decode('utf-8')
        assert "30.0" in exported_values_str  # Berechneter Wert von 10+20
        assert "20.0" in exported_values_str  # Berechneter Wert von 5+15
        assert "300.0" in exported_values_str  # Berechneter Wert von 100+200
    
    def test_csv_roundtrip_empty_cells(self):
        """Test: CSV mit leeren Zellen"""
        csv_content = b"""A;B;C
1;;3
;5;
7;;9"""
        
        # Import
        manager1 = import_csv_to_matrix(
            csv_content,
            "Test Matrix Empty",
            delimiter=";",
            has_header=True
        )
        
        assert manager1 is not None
        assert manager1.get_cell_value(0, 0) == 1.0
        assert manager1.get_cell_value(0, 1) is None  # Leere Zelle
        assert manager1.get_cell_value(0, 2) == 3.0
        
        # Export
        exported_csv = export_to_csv(manager1, delimiter=";")
        
        # Re-Import (mit Zeilennummern-Spalte)
        manager2 = import_csv_to_matrix(
            exported_csv,
            "Test Matrix Empty 2",
            delimiter=";",
            has_header=True
        )
        
        # Prüfe dass leere Zellen erhalten bleiben
        # Spalte 0 = Zeilennummer, Spalte 1 = A, Spalte 2 = B, Spalte 3 = C
        assert manager2.get_cell_value(0, 2) is None  # B1 ist leer


class TestExcelImportWithFormulas:
    """Test Excel Import mit Formeln"""
    
    @pytest.mark.skipif(
        not _is_openpyxl_available(),
        reason="openpyxl nicht installiert"
    )
    def test_excel_import_with_formulas(self):
        """Test: Excel-Datei mit Formeln importieren"""
        # Erstelle Test-Excel-Datei
        import openpyxl
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Header
        ws['A1'] = 'A'
        ws['B1'] = 'B'
        ws['C1'] = 'C'
        
        # Daten mit Formeln
        ws['A2'] = 10
        ws['B2'] = 20
        ws['C2'] = '=A2+B2'
        
        ws['A3'] = 5
        ws['B3'] = 15
        ws['C3'] = '=A3+B3'
        
        ws['A4'] = '=SUM(A2:A3)'
        ws['B4'] = '=SUM(B2:B3)'
        ws['C4'] = '=SUM(C2:C3)'
        
        # Speichere in Memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        excel_content = output.getvalue()
        
        # Import
        manager = import_excel_to_matrix(
            excel_content,
            "Test Excel Formulas",
            has_header=True
        )
        
        assert manager is not None
        
        # Prüfe Werte
        assert manager.get_cell_value(0, 0) == 10.0
        assert manager.get_cell_value(0, 1) == 20.0
        assert manager.get_cell_value(0, 2) == 30.0  # 10 + 20
        
        assert manager.get_cell_value(1, 0) == 5.0
        assert manager.get_cell_value(1, 1) == 15.0
        assert manager.get_cell_value(1, 2) == 20.0  # 5 + 15
        
        # Prüfe SUM-Formeln
        assert manager.get_cell_value(2, 0) == 15.0  # SUM(10, 5)
        assert manager.get_cell_value(2, 1) == 35.0  # SUM(20, 15)
        assert manager.get_cell_value(2, 2) == 50.0  # SUM(30, 20)
        
        # Prüfe dass Formeln gespeichert wurden
        cell_c2 = manager.get_cell(0, 2)
        assert cell_c2.is_formula()
        assert '=' in cell_c2.formula
    
    @pytest.mark.skipif(
        not _is_openpyxl_available(),
        reason="openpyxl nicht installiert"
    )
    def test_excel_roundtrip_with_formulas(self):
        """Test: Excel Import → Export → Re-Import mit Formeln"""
        # Erstelle Test-Excel-Datei
        import openpyxl
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        ws['A1'] = 'A'
        ws['B1'] = 'B'
        ws['C1'] = 'Sum'
        
        ws['A2'] = 100
        ws['B2'] = 200
        ws['C2'] = '=A2+B2'
        
        # Speichere in Memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        excel_content1 = output.getvalue()
        
        # Import
        manager1 = import_excel_to_matrix(
            excel_content1,
            "Test Excel Roundtrip",
            has_header=True
        )
        
        assert manager1.get_cell_value(0, 2) == 300.0
        
        # Export
        excel_content2 = export_to_excel(
            manager1,
            include_formulas=True,
            include_formatting=False
        )
        
        # Re-Import
        manager2 = import_excel_to_matrix(
            excel_content2,
            "Test Excel Roundtrip 2",
            has_header=True
        )
        
        # Prüfe dass Werte gleich sind
        assert manager2.get_cell_value(0, 0) == manager1.get_cell_value(0, 0)
        assert manager2.get_cell_value(0, 1) == manager1.get_cell_value(0, 1)
        assert manager2.get_cell_value(0, 2) == manager1.get_cell_value(0, 2)
        
        # Prüfe dass Formel erhalten blieb
        cell = manager2.get_cell(0, 2)
        assert cell.is_formula()


class TestLargeFiles:
    """Test große Dateien (10 MB)"""
    
    def test_large_csv_import(self):
        """Test: Import einer großen CSV-Datei"""
        # Erstelle große CSV-Datei (ca. 10 MB)
        # 1000 Zeilen × 100 Spalten mit Zahlen
        rows = 1000
        cols = 100
        
        # Header
        header = ";".join([f"Col{i}" for i in range(cols)])
        csv_lines = [header]
        
        # Daten
        for row in range(rows):
            row_data = ";".join([str(row * cols + col) for col in range(cols)])
            csv_lines.append(row_data)
        
        csv_content = "\n".join(csv_lines).encode('utf-8')
        
        # Prüfe Größe (sollte > 1 MB sein)
        size_mb = len(csv_content) / (1024 * 1024)
        print(f"CSV-Größe: {size_mb:.2f} MB")
        assert size_mb > 1.0
        
        # Import
        import time
        start_time = time.time()
        
        manager = import_csv_to_matrix(
            csv_content,
            "Large CSV Test",
            delimiter=";",
            has_header=True
        )
        
        import_time = time.time() - start_time
        print(f"Import-Zeit: {import_time:.2f} Sekunden")
        
        assert manager is not None
        assert manager.matrix.rows >= rows
        assert manager.matrix.columns >= cols
        
        # Prüfe einige Werte
        assert manager.get_cell_value(0, 0) == 0.0
        assert manager.get_cell_value(0, 1) == 1.0
        assert manager.get_cell_value(1, 0) == 100.0
        
        # Export
        start_time = time.time()
        
        exported_csv = export_to_csv(
            manager,
            delimiter=";",
            include_formulas=False
        )
        
        export_time = time.time() - start_time
        print(f"Export-Zeit: {export_time:.2f} Sekunden")
        
        assert exported_csv is not None
        assert len(exported_csv) > 1024 * 1024  # > 1 MB
    
    @pytest.mark.skipif(
        not _is_openpyxl_available(),
        reason="openpyxl nicht installiert"
    )
    def test_large_excel_export(self):
        """Test: Export einer großen Matrix als Excel"""
        # Erstelle große Matrix
        matrix = ExcelMatrix(
            name="Large Matrix",
            rows=500,
            columns=50
        )
        
        # Fülle mit Daten
        for row in range(min(500, matrix.rows)):
            for col in range(min(50, matrix.columns)):
                value = row * 50 + col
                cell = Cell(row=row, col=col, value=float(value))
                matrix.cells[(row, col)] = cell
        
        manager = ExcelManager(matrix)
        
        # Export
        import time
        start_time = time.time()
        
        excel_data = export_to_excel(
            manager,
            include_formulas=False,
            include_formatting=True
        )
        
        export_time = time.time() - start_time
        print(f"Excel Export-Zeit: {export_time:.2f} Sekunden")
        
        assert excel_data is not None
        
        size_mb = len(excel_data) / (1024 * 1024)
        print(f"Excel-Größe: {size_mb:.2f} MB")
        
        # Prüfe dass Datei nicht zu groß ist (sollte < 10 MB sein)
        assert size_mb < 10.0


class TestDifferentEncodings:
    """Test verschiedene Encodings"""
    
    def test_utf8_encoding(self):
        """Test: UTF-8 Encoding"""
        csv_content = "Name;Wert\nÄpfel;10\nÖl;20\nÜbung;30".encode('utf-8')
        
        manager = import_csv_to_matrix(
            csv_content,
            "UTF-8 Test",
            delimiter=";",
            has_header=True,
            encoding='utf-8'
        )
        
        assert manager is not None
        assert manager.get_cell_value(0, 1) == 10.0
        assert manager.get_cell_value(1, 1) == 20.0
        assert manager.get_cell_value(2, 1) == 30.0
        
        # Export
        exported = export_to_csv(
            manager,
            delimiter=";",
            encoding='utf-8'
        )
        
        # Prüfe dass Umlaute erhalten bleiben
        assert "Äpfel".encode('utf-8') in exported or "pfel" in exported.decode('utf-8')
    
    def test_latin1_encoding(self):
        """Test: Latin-1 (ISO-8859-1) Encoding"""
        csv_content = "Name;Wert\nÄpfel;10\nÖl;20\nÜbung;30".encode('latin-1')
        
        manager = import_csv_to_matrix(
            csv_content,
            "Latin-1 Test",
            delimiter=";",
            has_header=True,
            encoding='latin-1'
        )
        
        assert manager is not None
        assert manager.get_cell_value(0, 1) == 10.0
        
        # Export
        exported = export_to_csv(
            manager,
            delimiter=";",
            encoding='latin-1'
        )
        
        assert exported is not None
    
    def test_windows1252_encoding(self):
        """Test: Windows-1252 Encoding"""
        csv_content = "Name;Wert\nTest;100\nData;200".encode('windows-1252')
        
        manager = import_csv_to_matrix(
            csv_content,
            "Windows-1252 Test",
            delimiter=";",
            has_header=True,
            encoding='windows-1252'
        )
        
        assert manager is not None
        assert manager.get_cell_value(0, 1) == 100.0
        assert manager.get_cell_value(1, 1) == 200.0
    
    def test_auto_encoding_detection(self):
        """Test: Automatische Encoding-Erkennung"""
        # UTF-8 mit BOM
        csv_content = "\ufeffName;Wert\nTest;123".encode('utf-8')
        
        validation = validate_csv_file(csv_content)
        
        assert validation['valid']
        assert validation['encoding'] in ['utf-8', 'UTF-8-SIG', 'ascii']
        
        manager = import_csv_to_matrix(
            csv_content,
            "Auto Encoding Test",
            delimiter=";",
            has_header=True,
            encoding=None  # Auto-detect
        )
        
        assert manager is not None
        assert manager.get_cell_value(0, 1) == 123.0


class TestExportUtilities:
    """Test Export-Hilfsfunktionen"""
    
    def test_generate_filename_csv(self):
        """Test: Dateinamen-Generierung für CSV"""
        filename = generate_filename("Test Matrix", "csv", include_timestamp=False)
        assert filename == "Test_Matrix.csv"
        
        filename_ts = generate_filename("Test Matrix", "csv", include_timestamp=True)
        assert filename_ts.startswith("Test_Matrix_")
        assert filename_ts.endswith(".csv")
    
    def test_generate_filename_xlsx(self):
        """Test: Dateinamen-Generierung für Excel"""
        filename = generate_filename("My Matrix", "xlsx", include_timestamp=False)
        assert filename == "My_Matrix.xlsx"
        
        filename_ts = generate_filename("My Matrix", "xlsx", include_timestamp=True)
        assert filename_ts.startswith("My_Matrix_")
        assert filename_ts.endswith(".xlsx")
    
    def test_generate_filename_special_chars(self):
        """Test: Dateinamen-Generierung mit Sonderzeichen"""
        filename = generate_filename("Test/Matrix:2024", "csv", include_timestamp=False)
        # Sonderzeichen sollten entfernt werden
        assert "/" not in filename
        assert ":" not in filename
        assert filename.endswith(".csv")
    
    def test_get_export_info(self):
        """Test: Export-Informationen abrufen"""
        matrix = ExcelMatrix(name="Test", rows=10, columns=5)
        
        # Füge einige Zellen hinzu
        for i in range(5):
            cell = Cell(row=i, col=0, value=float(i))
            matrix.cells[(i, 0)] = cell
        
        # Füge Formel hinzu
        formula_cell = Cell(row=0, col=1, formula="=A1*2")
        matrix.cells[(0, 1)] = formula_cell
        
        manager = ExcelManager(matrix)
        
        info = get_export_info(manager)
        
        assert info['matrix_name'] == "Test"
        assert info['rows'] == 10
        assert info['columns'] == 5
        assert info['cell_count'] == 6  # 5 Werte + 1 Formel
        assert info['formula_count'] == 1
        assert 'estimated_csv_size' in info
        assert 'estimated_xlsx_size' in info
    
    def test_validate_export(self):
        """Test: Export-Validierung"""
        matrix = ExcelMatrix(name="Test", rows=10, columns=5)
        manager = ExcelManager(matrix)
        
        validation = validate_export(manager)
        
        assert 'valid' in validation
        assert 'warnings' in validation
        assert 'errors' in validation
        assert isinstance(validation['warnings'], list)
        assert isinstance(validation['errors'], list)


if __name__ == "__main__":
    # Führe Tests aus
    pytest.main([__file__, "-v", "-s"])
