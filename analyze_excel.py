"""Analysiere alle Excel-Dateien im wp_implements/excel Ordner"""
import openpyxl
import os

excel_dir = "wp_implements/excel"
files = ['1.xlsx', '2.xlsx', '3.xlsx', '4.xlsx', '5.xlsx']

for filename in files:
    filepath = os.path.join(excel_dir, filename)
    print(f"\n{'='*80}")
    print(f"FILE: {filename}")
    print('='*80)
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        print(f"Sheets: {wb.sheetnames}\n")
        
        for sheet in wb.worksheets:
            print(f"\n--- SHEET: {sheet.title} ---")
            print(f"Dimensions: {sheet.dimensions}")
            print("\nFirst 30 rows:")
            
            for i, row in enumerate(sheet.iter_rows(max_row=30, values_only=True), 1):
                # Filter out empty rows
                if any(cell is not None for cell in row):
                    print(f"Row {i}: {row}")
            
    except Exception as e:
        print(f"Error reading {filename}: {e}")
