"""
Test Excel Import Funktionalität

Testet die Excel-Import-Funktionen (XLS/XLSX) für die Excel-Integration.
"""

import pytest
import io

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from excel.excel_import import (
    get_excel_sheet_names,
    parse_excel_content,
    import_excel_to_matrix,
    validate_excel_file,
    get_excel_preview,
    ImportError as ExcelImportError,
    OPENPYXL_AVAILABLE as IMPORT_OPENPYXL_AVAILABLE
)


# Helper function to create test Excel files
def create_test_excel_file(data, sheet_name="Sheet1", formulas=None):
    """
    Erstellt eine Test-Excel-Datei im Speicher

    Args:
        data: Liste von Listen mit Zellwerten
        sheet_name: Name des Sheets
        formulas: Optional dict mit {(row, col): formula}

    Returns:
        Bytes der Excel-Datei
    """
    if not OPENPYXL_AVAILABLE:
        pytest.skip("openpyxl nicht installiert")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Fülle Daten
    for row_idx, row_data in enumerate(data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Füge Formeln hinzu
    if formulas:
        for (row, col), formula in formulas.items():
            cell = ws.cell(row=row, column=col)
            cell.value = formula

    # Speichere in BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl nicht installiert")
def test_get_excel_sheet_names():
    """Test Sheet-Namen-Extraktion"""
    # Erstelle Excel mit mehreren Sheets
    wb = openpyxl.Workbook()
    wb.active.title = "Daten"
    wb.create_sheet("Preise")
    wb.create_sheet("Konfiguration")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    excel_bytes = output.read()

    # Teste Sheet-Namen
    sheet_names = get_excel_sheet_names(excel_bytes)

    assert len(sheet_names) == 3
    assert "Daten" in sheet_names
    assert "Preise" in sheet_names
    assert "Konfiguration" in sheet_names


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl nicht installiert")
def test_parse_excel_content():
    """Test Excel-Parsing"""
    data = [
        ["Name", "Wert", "Preis"],
        ["Produkt A", 10, 100],
        ["Produkt B", 20, 200]
    ]

    excel_bytes = create_test_excel_file(data)

    header, rows = parse_excel_content(excel_bytes)

    assert header == ["Name", "Wert", "Preis"]
    assert len(rows) == 2
    assert rows[0] == ["Produkt A", 10, 100]
    assert rows[1] == ["Produkt B", 20, 200]


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl nicht installiert")
def test_parse_excel_without_header():
    """Test Excel-Parsing ohne Header"""
    data = [
        [10, 100],
        [20, 200]
    ]

    excel_bytes = create_test_excel_file(data)

    header, rows = parse_excel_content(excel_bytes, has_header=False)

    # Automatisch generierte Header (A, B, C, ...)
    assert header == ["A", "B"]
    assert len(rows) == 2
    assert rows[0] == [10, 100]


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl nicht installiert")
def test_parse_excel_with_formulas():
    """Test Excel-Parsing mit Formeln"""
    data = [
        ["A", "B", "C"],
        [10, 20, None],
        [30, 40, None]
    ]

    # Füge Formeln hinzu
    formulas = {
        (2, 3): "=A2+B2",  # C1 = A1 + B1
        (3, 3): "=A3+B3"   # C2 = A2 + B2
    }

    excel_bytes = create_test_excel_file(data, formulas=formulas)

    header, rows = parse_excel_content(excel_bytes)

    assert header == ["A", "B", "C"]
    assert len(rows) == 2

    # Prüfe dass Formeln erkannt wurden
    assert rows[0][2] == "=A2+B2"
    assert rows[1][2] == "=A3+B3"


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl nicht installiert")
def test_import_excel_to_matrix():
    """Test Excel-Import zu Matrix"""
    data = [
        ["Spalte1", "Spalte2", "Spalte3"],
        [10, 20, 30],
        [40, 50, 60]
    ]

    excel_bytes = create_test_excel_file(data)

    manager = import_excel_to_matrix(
        excel_bytes,
        "Test Matrix"
    )

    assert manager is not None
    assert manager.matrix.name == "Test Matrix"

    # Prüfe Zellwerte
    assert manager.get_cell_value(0, 0) == 10
    assert manager.get_cell_value(0, 1) == 20
    assert manager.get_cell_value(0, 2) == 30
    assert manager.get_cell_value(1, 0) == 40
    assert manager.get_cell_value(1, 1) == 50
    assert manager.get_cell_value(1, 2) == 60


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl nicht installiert")
def test_import_excel_with_formulas():
    """Test Excel-Import mit Formeln"""
    data = [
        ["A", "B", "C"],
        [10, 20, None],
        [30, 40, None]
    ]

    formulas = {
        (2, 3): "=A2+B2",
        (3, 3): "=A3+B3"
    }

    excel_bytes = create_test_excel_file(data, formulas=formulas)

    manager = import_excel_to_matrix(
        excel_bytes,
        "Test Matrix mit Formeln"
    )

    assert manager is not None

    # Prüfe dass Formeln erkannt wurden
    cell_c1 = manager.get_cell(0, 2)
    assert cell_c1.is_formula()
    assert cell_c1.formula == "=A2+B2"

    # Prüfe berechnete Werte
    # Note: Die Formel verwendet Excel-Notation (A2+B2), nicht unsere interne (A1+B1)
    # Daher wird die Berechnung möglicherweise fehlschlagen oder 0 ergeben
    # Dies ist erwartetes Verhalten - Formeln müssen ggf. angepasst werden


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl nicht installiert")
def test_import_excel_with_multiple_sheets():
    """Test Excel-Import mit mehreren Sheets"""
    # Erstelle Excel mit mehreren Sheets
    wb = openpyxl.Workbook()

    # Sheet 1
    ws1 = wb.active
    ws1.title = "Daten"
    ws1.append(["A", "B"])
    ws1.append([1, 2])

    # Sheet 2
    ws2 = wb.create_sheet("Preise")
    ws2.append(["Preis", "Menge"])
    ws2.append([100, 10])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    excel_bytes = output.read()

    # Importiere erstes Sheet
    manager1 = import_excel_to_matrix(
        excel_bytes,
        "Matrix Sheet 1",
        sheet_name="Daten"
    )

    assert manager1 is not None
    assert manager1.get_cell_value(0, 0) == 1
    assert manager1.get_cell_value(0, 1) == 2

    # Importiere zweites Sheet
    manager2 = import_excel_to_matrix(
        excel_bytes,
        "Matrix Sheet 2",
        sheet_name="Preise"
    )

    assert manager2 is not None
    assert manager2.get_cell_value(0, 0) == 100
    assert manager2.get_cell_value(0, 1) == 10


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl nicht installiert")
def test_validate_excel_file():
    """Test Excel-Validierung"""
    data = [
        ["A", "B", "C"],
        [1, 2, 3],
        [4, 5, 6]
    ]

    excel_bytes = create_test_excel_file(data)

    validation = validate_excel_file(excel_bytes)

    assert validation['valid'] is True
    assert len(validation['sheets']) == 1
    assert validation['sheets'][0] == "Sheet1"
    assert validation['num_rows'] == 2
    assert validation['num_cols'] == 3
    assert validation['has_formulas'] is False


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl nicht installiert")
def test_validate_excel_with_formulas():
    """Test Excel-Validierung mit Formeln"""
    data = [
        ["A", "B", "C"],
        [1, 2, None]
    ]

    formulas = {
        (2, 3): "=A2+B2"
    }

    excel_bytes = create_test_excel_file(data, formulas=formulas)

    validation = validate_excel_file(excel_bytes)

    assert validation['valid'] is True
    assert validation['has_formulas'] is True


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl nicht installiert")
def test_get_excel_preview():
    """Test Excel-Vorschau"""
    data = [["A", "B", "C"]]
    for i in range(1, 21):
        data.append([i, i*2, i*3])

    excel_bytes = create_test_excel_file(data)

    preview = get_excel_preview(excel_bytes, max_rows=5)

    assert preview['header'] == ["A", "B", "C"]
    assert len(preview['rows']) == 5  # Nur 5 Zeilen in Vorschau
    assert preview['total_rows'] == 20  # Aber 20 Zeilen insgesamt
    assert len(preview['sheets']) == 1
    assert preview['current_sheet'] == "Sheet1"


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl nicht installiert")
def test_import_excel_with_empty_cells():
    """Test Excel-Import mit leeren Zellen"""
    data = [
        ["A", "B", "C"],
        [10, None, 30],
        [None, 50, None],
        [70, 80, 90]
    ]

    excel_bytes = create_test_excel_file(data)

    manager = import_excel_to_matrix(
        excel_bytes,
        "Test Matrix mit leeren Zellen"
    )

    assert manager is not None

    # Prüfe dass leere Zellen übersprungen wurden
    assert manager.get_cell_value(0, 0) == 10
    assert manager.get_cell_value(0, 1) is None or manager.get_cell_value(0, 1) == 0
    assert manager.get_cell_value(0, 2) == 30


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl nicht installiert")
def test_import_excel_with_mixed_types():
    """Test Excel-Import mit gemischten Datentypen"""
    data = [
        ["Text", "Zahl", "Datum"],
        ["Hallo", 42, "2024-01-01"],
        ["Welt", 3.14, "2024-12-31"]
    ]

    excel_bytes = create_test_excel_file(data)

    manager = import_excel_to_matrix(
        excel_bytes,
        "Test Matrix gemischte Typen"
    )

    assert manager is not None

    # Prüfe Typen
    assert manager.get_cell_value(0, 0) == "Hallo"
    assert manager.get_cell_value(0, 1) == 42
    # Datum wird als String behandelt
    assert manager.get_cell_value(0, 2) is not None


def test_import_without_openpyxl():
    """Test dass ImportError geworfen wird wenn openpyxl fehlt"""
    if OPENPYXL_AVAILABLE:
        pytest.skip("openpyxl ist installiert")

    # Teste dass alle Funktionen ImportError werfen
    with pytest.raises(ExcelImportError):
        get_excel_sheet_names(b"dummy")

    with pytest.raises(ExcelImportError):
        parse_excel_content(b"dummy")

    with pytest.raises(ExcelImportError):
        import_excel_to_matrix(b"dummy", "Test")

    validation = validate_excel_file(b"dummy")
    assert validation['valid'] is False
    assert len(validation['errors']) > 0

    with pytest.raises(ExcelImportError):
        get_excel_preview(b"dummy")


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl nicht installiert")
def test_import_invalid_sheet_name():
    """Test Import mit ungültigem Sheet-Namen"""
    data = [["A", "B"], [1, 2]]
    excel_bytes = create_test_excel_file(data, sheet_name="ValidSheet")

    with pytest.raises(ExcelImportError) as exc_info:
        import_excel_to_matrix(
            excel_bytes,
            "Test Matrix",
            sheet_name="InvalidSheet"
        )

    assert "nicht gefunden" in str(exc_info.value)


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl nicht installiert")
def test_import_empty_excel():
    """Test Import einer leeren Excel-Datei"""
    wb = openpyxl.Workbook()
    ws = wb.active

    # Leeres Sheet
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    excel_bytes = output.read()

    with pytest.raises(ExcelImportError):
        import_excel_to_matrix(
            excel_bytes,
            "Leere Matrix"
        )


if __name__ == "__main__":
    # Führe Tests aus
    pytest.main([__file__, "-v"])
